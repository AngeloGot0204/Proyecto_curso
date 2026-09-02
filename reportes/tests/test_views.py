"""Integration tests for the wizard-captura views (backlog #5, Phase 4;
design D7, D8, D9; specs `wizard-captura` and `reportes-modelo`).

`test_post_nuevo_*` / `test_get_nuevo_*` cover design D7 (`POST
/reportes/<codigo>/nuevo/` creates the `Reporte`, `require_POST`,
`login_required`). `test_*_paso_*` cover GET rehydration, POST upsert (no
duplicate rows on re-POST), the creator-only 404 (D9), the unknown-
`seccion_id` 404, and the non-blocking `obligatorio` marker (spec:
Non-blocking obligatorio marker).
"""

import re
from io import BytesIO
from unittest import mock

import pytest
from django.contrib.messages import get_messages
from django.core.files.uploadedfile import SimpleUploadedFile
from django.template.loader import render_to_string
from django.urls import reverse
from django.utils import formats
from openpyxl import load_workbook

from reportes.models import (
    CambioDeValor,
    EstadoDeReporte,
    Generacion,
    ParticipacionEnReporte,
    Reporte,
    ValorDeReporte,
    VistoBueno,
)
from reportes.validacion import validar_reporte
from tipos_reporte.generador import PlantillaIlegible
from tipos_reporte.models import TipoDeReporte


@pytest.fixture
def sesion_de_creador(client, usuario_factory, reporte_factory):
    """A `(client, reporte)` pair where the logged-in user IS the `Reporte`'s
    `creador` — needed for every test exercising the happy path through
    `paso`, since `reporte_factory`'s default creador is otherwise an
    unrelated, freshly-made usuario."""
    creador = usuario_factory(username="creador_del_reporte")
    reporte = reporte_factory(creador=creador)
    client.force_login(creador)
    return client, reporte


@pytest.fixture
def sesion_de_invitado(client, reporte_factory, participacion_factory):
    """A `(client, reporte)` pair where the logged-in user is an INVITED
    non-creator participant (backlog #8, tasks.md 3.16) — mirrors
    `sesion_de_creador`, but the logged-in user has a `ParticipacionEnReporte`
    row instead of being `Reporte.creador`."""
    reporte = reporte_factory()
    invitado = participacion_factory(reporte, username="invitado_de_sesion")
    client.force_login(invitado)
    return client, reporte


# ---------------------------------------------------------------------------
# iniciar_reporte
# ---------------------------------------------------------------------------


@pytest.mark.django_db
def test_post_nuevo_crea_un_reporte(cliente_autenticado, tipo_con_definicion_activa_factory):
    tipo, definicion = tipo_con_definicion_activa_factory()

    response = cliente_autenticado.post(
        reverse("reportes_nuevo", args=[tipo.codigo])
    )

    assert Reporte.objects.count() == 1
    reporte = Reporte.objects.get()
    assert reporte.tipo_id == tipo.id
    assert reporte.definicion_id == definicion.id
    assert response.status_code == 302
    primera_seccion = definicion.estructura["secciones"][0]["id"]
    assert response.url == reverse(
        "reportes_paso", args=[reporte.id, primera_seccion]
    )


@pytest.mark.django_db
def test_get_nuevo_no_permitido(cliente_autenticado, tipo_con_definicion_activa_factory):
    tipo, _definicion = tipo_con_definicion_activa_factory()

    response = cliente_autenticado.get(reverse("reportes_nuevo", args=[tipo.codigo]))

    assert response.status_code == 405
    assert Reporte.objects.count() == 0


@pytest.mark.django_db
def test_nuevo_anonimo_redirige_a_login(client, tipo_con_definicion_activa_factory):
    tipo, _definicion = tipo_con_definicion_activa_factory()

    response = client.post(reverse("reportes_nuevo", args=[tipo.codigo]))

    assert response.status_code == 302
    assert reverse("login") in response.url
    assert Reporte.objects.count() == 0


# ---------------------------------------------------------------------------
# paso — GET rehydration
# ---------------------------------------------------------------------------


@pytest.mark.django_db
def test_get_paso_rehidrata_valores_guardados(sesion_de_creador):
    client, reporte = sesion_de_creador
    ValorDeReporte.objects.create(
        reporte=reporte,
        identificador_de_campo="turno",
        valor="Día",
        autor=reporte.creador,
    )

    response = client.get(reverse("reportes_paso", args=[reporte.id, "datos-generales"]))

    assert response.status_code == 200
    assert response.context["form"].initial.get("turno") == "Día"


@pytest.mark.django_db
def test_paso_anonimo_redirige_a_login(client, reporte_factory):
    reporte = reporte_factory()

    response = client.get(
        reverse("reportes_paso", args=[reporte.id, "datos-generales"])
    )

    assert response.status_code == 302
    assert reverse("login") in response.url


@pytest.mark.django_db
def test_paso_reporte_de_otro_usuario_da_404(
    cliente_autenticado, usuario_factory, reporte_factory
):
    # A distinct usuario than the one `cliente_autenticado` logged in
    # (design D9) — `usuario_factory`'s default username collides with
    # `cliente_autenticado`'s own default user otherwise.
    otro_creador = usuario_factory(username="otro_usuario")
    reporte = reporte_factory(creador=otro_creador)

    response = cliente_autenticado.get(
        reverse("reportes_paso", args=[reporte.id, "datos-generales"])
    )

    assert response.status_code == 404


@pytest.mark.django_db
def test_paso_participante_invitado_accede(sesion_de_invitado):
    """Spec `wizard-captura` — "Invited participant accesses a step": an
    invited non-creator user gets 200, not 404 (backlog #8, tasks.md 3.1)."""
    client, reporte = sesion_de_invitado

    response = client.get(
        reverse("reportes_paso", args=[reporte.id, "datos-generales"])
    )

    assert response.status_code == 200


@pytest.mark.django_db
def test_paso_no_invitado_autenticado_da_404(
    cliente_autenticado, usuario_factory, reporte_factory
):
    """Spec `wizard-captura` — "Non-invited authenticated user is denied":
    a user with no `ParticipacionEnReporte` row and who is not the creator
    still gets 404 via a direct URL (backlog #8, tasks.md 3.2)."""
    otro_creador = usuario_factory(username="creador_paso_no_invitado")
    reporte = reporte_factory(creador=otro_creador)

    response = cliente_autenticado.get(
        reverse("reportes_paso", args=[reporte.id, "datos-generales"])
    )

    assert response.status_code == 404


@pytest.mark.django_db
def test_paso_seccion_desconocida_da_404(sesion_de_creador):
    client, reporte = sesion_de_creador

    response = client.get(
        reverse("reportes_paso", args=[reporte.id, "seccion-inexistente"])
    )

    assert response.status_code == 404


@pytest.mark.django_db
def test_get_paso_seccion_vacia_renderiza_sin_error(
    client, usuario_factory, tipo_con_definicion_activa_factory
):
    """Spec: "Section with no campos/items still renders" — the step shell
    must render without error, showing zero fields, and still expose next-
    step navigation."""
    estructura = {
        "secciones": [
            {"id": "vacia", "titulo": "Sección vacía", "campos": []},
            {"id": "siguiente", "titulo": "Siguiente", "campos": []},
        ]
    }
    tipo, definicion = tipo_con_definicion_activa_factory(estructura=estructura)
    creador = usuario_factory(username="creador_seccion_vacia")
    client.force_login(creador)

    from reportes.models import Reporte

    reporte = Reporte.objects.create(
        tipo=tipo, definicion=definicion, creador=creador
    )

    response = client.get(reverse("reportes_paso", args=[reporte.id, "vacia"]))

    assert response.status_code == 200
    assert list(response.context["form"].fields) == []
    assert response.context["url_siguiente"] == reverse(
        "reportes_paso", args=[reporte.id, "siguiente"]
    )


# ---------------------------------------------------------------------------
# paso — POST upsert
# ---------------------------------------------------------------------------


@pytest.mark.django_db
def test_post_paso_persiste_valores(sesion_de_creador):
    client, reporte = sesion_de_creador

    response = client.post(
        reverse("reportes_paso", args=[reporte.id, "datos-generales"]),
        data={"turno": "Noche"},
    )

    assert response.status_code == 302
    valor = ValorDeReporte.objects.get(reporte=reporte, identificador_de_campo="turno")
    assert valor.valor == "Noche"


@pytest.mark.django_db
def test_post_paso_repetido_no_duplica_fila(sesion_de_creador):
    client, reporte = sesion_de_creador

    client.post(
        reverse("reportes_paso", args=[reporte.id, "datos-generales"]),
        data={"turno": "Día"},
    )
    client.post(
        reverse("reportes_paso", args=[reporte.id, "datos-generales"]),
        data={"turno": "Noche"},
    )

    assert (
        ValorDeReporte.objects.filter(
            reporte=reporte, identificador_de_campo="turno"
        ).count()
        == 1
    )
    valor = ValorDeReporte.objects.get(reporte=reporte, identificador_de_campo="turno")
    assert valor.valor == "Noche"


@pytest.mark.django_db
def test_post_paso_rango_hora_inicio_fin_persiste_dos_filas(sesion_de_creador):
    client, reporte = sesion_de_creador

    response = client.post(
        reverse("reportes_paso", args=[reporte.id, "proceso-instalacion"]),
        data={"p-01_inicio": "08:00", "p-01_fin": "10:00"},
    )

    assert response.status_code == 302
    inicio = ValorDeReporte.objects.get(
        reporte=reporte, identificador_de_campo="p-01_inicio"
    )
    fin = ValorDeReporte.objects.get(
        reporte=reporte, identificador_de_campo="p-01_fin"
    )
    assert inicio.valor == "08:00"
    assert fin.valor == "10:00"


@pytest.mark.django_db
def test_post_paso_ultima_seccion_redirige_a_si_misma(sesion_de_creador):
    client, reporte = sesion_de_creador

    response = client.post(
        reverse("reportes_paso", args=[reporte.id, "proceso-instalacion"]),
        data={},
    )

    assert response.status_code == 302
    assert response.url == reverse(
        "reportes_paso", args=[reporte.id, "proceso-instalacion"]
    )


@pytest.mark.django_db
def test_post_paso_no_ultima_seccion_redirige_a_siguiente(sesion_de_creador):
    client, reporte = sesion_de_creador

    response = client.post(
        reverse("reportes_paso", args=[reporte.id, "datos-generales"]),
        data={"turno": "Día"},
    )

    assert response.status_code == 302
    assert response.url == reverse(
        "reportes_paso", args=[reporte.id, "proceso-instalacion"]
    )


@pytest.mark.django_db
def test_paso_post_redirect_es_seguible(sesion_de_creador):
    """Change `sincronizacion-numero-registro` (design's Interfaces/D4;
    spec `reporte-idempotent-creation` scenario 11): `paso`'s POST answers
    302, and following it with `follow=True` lands 200 on the next step —
    exactly the `response.redirected`/`response.url` contract the
    fetch-based client submit depends on."""
    client, reporte = sesion_de_creador

    response = client.post(
        reverse("reportes_paso", args=[reporte.id, "datos-generales"]),
        data={"turno": "Día"},
        follow=True,
    )

    assert response.redirect_chain
    assert response.status_code == 200


# ---------------------------------------------------------------------------
# paso — client-side JS contract (backlog validacion-datos-formulario, Phase 5)
# ---------------------------------------------------------------------------


@pytest.mark.django_db
def test_get_paso_incluye_atributos_data_y_script_paso_js(
    client, estructura_con_validaciones, reporte_con_validaciones_factory
):
    """GET `paso` HTML must carry the JS contract's data attributes plus a
    deferred `paso.js` script tag (spec: "Client-side hora range feedback",
    "\"No cumple\" observación toggling"; design's `paso.html`/`paso.js`
    File Changes rows and "the JS contract" subsection). No JS test runner
    exists in this project, so this rendered-attribute contract is the only
    coverage for `paso.js` (design's Testing Strategy, explicitly
    accepted)."""
    reporte = reporte_con_validaciones_factory(client, estructura_con_validaciones)

    respuesta_datos = client.get(
        reverse("reportes_paso", args=[reporte.id, "datos-generales"])
    )
    contenido_datos = respuesta_datos.content.decode()

    assert respuesta_datos.status_code == 200
    assert 'data-campo="observaciones-generales"' in contenido_datos
    assert 'data-campo="estado-general"' in contenido_datos
    assert 'data-campo="estado-general_observacion"' in contenido_datos
    assert 'data-requiere-observacion="estado-general_observacion"' in contenido_datos
    assert "data-siguiente" in contenido_datos
    assert 'reportes/paso.js"' in contenido_datos
    assert " defer" in contenido_datos

    respuesta_proceso = client.get(
        reverse("reportes_paso", args=[reporte.id, "proceso-instalacion"])
    )
    contenido_proceso = respuesta_proceso.content.decode()

    assert respuesta_proceso.status_code == 200
    assert 'data-campo="p-01_inicio"' in contenido_proceso
    assert 'data-campo="p-01_fin"' in contenido_proceso
    assert 'data-rango="p-01"' in contenido_proceso


@pytest.mark.django_db
def test_paso_aplica_pasos_indicador_y_checklist_component_classes(sesion_de_creador):
    """Change `retrofit-visual-design2` PR2 (design D3, task 3.10): the
    wizard step indicator uses `.pasos` (DESIGN2 §4 "Indicador de pasos")
    and the field list uses `.checklist` (DESIGN2 §4 "Checklist por rol",
    reasonable extrapolation onto the generic `paso.html` field loop, since
    no per-role data structure exists in this app — spec `visual-design-
    system`, requirement 'Eight DESIGN2 Component Classes')."""
    client, reporte = sesion_de_creador

    response = client.get(
        reverse("reportes_paso", args=[reporte.id, "datos-generales"])
    )
    contenido = response.content.decode()

    assert response.status_code == 200
    assert 'class="pasos' in contenido
    assert 'class="checklist' in contenido
    assert 'checklist__item' in contenido


@pytest.mark.django_db
def test_paso_fila_de_horas_aplica_component_class(
    client, estructura_con_validaciones, reporte_con_validaciones_factory
):
    """Change `retrofit-visual-design2` PR2 (design D3/§6.b, task 3.10): a
    `rango-hora-inicio-fin` item's two fields (`_inicio`/`_fin` suffix,
    `tipos_reporte.generador._SUFIJO_POR_CLAVE`) both get the `.fila-horas`
    modifier so they lay out side-by-side (DESIGN2 §6.b "Fila de horas")."""
    reporte = reporte_con_validaciones_factory(client, estructura_con_validaciones)

    response = client.get(
        reverse("reportes_paso", args=[reporte.id, "proceso-instalacion"])
    )
    contenido = response.content.decode()

    assert response.status_code == 200
    assert contenido.count("fila-horas") == 2


@pytest.mark.django_db
def test_paso_primer_boton_submit_sigue_siendo_el_del_formulario_tras_retrofit(
    sesion_de_creador,
):
    """Submit-target guard for `paso.js`'s `aplicarEstadoDeNavegacion`.

    This test used to assert that the FIRST `<button type="submit">` on the
    whole page was the step's own — a guard that matched `paso.js`'s
    unscoped `querySelector('form button[type="submit"]')`. That assumption
    stopped holding when `db1c1d6` gave the sidebar a logout `<form>`, which
    renders before the page content: the selector started resolving to
    "Cerrar sesión", so an invalid time range disabled the LOGOUT button and
    left "Guardar y continuar" enabled — the validation guard silently
    stopped guarding.

    The fix scoped the selector to the step form. So the property worth
    pinning is no longer "first on the page" (a DOM-order coincidence any
    layout change can break) but "resolves to the step's own submit button",
    which is what the code actually needs. Asserted on both halves of that
    contract: the markup exposes the step form under the same
    `[data-reporte-id][data-seccion-id]` attributes `paso-offline.js`
    already keys on, and `paso.js` scopes its lookup to it."""
    from pathlib import Path

    client, reporte = sesion_de_creador

    response = client.get(
        reverse("reportes_paso", args=[reporte.id, "datos-generales"])
    )
    contenido = response.content.decode()

    assert response.status_code == 200

    # Half 1 — the markup contract the selector depends on.
    formulario = re.search(
        r"<form[^>]*data-reporte-id[^>]*data-seccion-id[^>]*>(.*?)</form>",
        contenido,
        re.DOTALL,
    )
    assert formulario is not None, "el paso no expone su form con los data-attrs"
    boton_del_paso = re.search(
        r'<button[^>]*type="submit"[^>]*>([^<]*)</button>', formulario.group(1)
    )
    assert boton_del_paso is not None
    assert "Guardar y continuar" in boton_del_paso.group(1)

    # Half 2 — the lookup stays scoped. An unscoped `form button[...]` would
    # pass half 1 and still grab the sidebar's logout button at runtime.
    paso_js = (
        Path(__file__).resolve().parents[1] / "static" / "reportes" / "paso.js"
    ).read_text(encoding="utf-8")
    assert 'document.querySelector(\'form button[type="submit"]\')' not in paso_js
    assert 'form[data-reporte-id][data-seccion-id] button[type="submit"]' in paso_js


@pytest.mark.django_db
def test_paso_carga_envio_paso_js_antes_que_paso_offline_js(sesion_de_creador):
    """Change `vista-sincronizacion-pendientes`, Phase 1 (design's D2) —
    `envio-paso.js` defines `window.reportesEnvioPaso`, which
    `paso-offline.js` now depends on at parse/execute time, so the shared
    helper MUST be present in the document and load before it."""
    client, reporte = sesion_de_creador

    response = client.get(
        reverse("reportes_paso", args=[reporte.id, "datos-generales"])
    )
    contenido = response.content.decode()

    assert response.status_code == 200
    assert "reportes/envio-paso.js" in contenido
    assert contenido.index("reportes/envio-paso.js") < contenido.index(
        "reportes/paso-offline.js"
    )


@pytest.mark.django_db
def test_paso_expone_tipo_nombre_y_fecha_reporte_en_data_attrs(sesion_de_creador):
    """Change `vista-sincronizacion-pendientes`, Phase 2 (design's D4) —
    the step `<form>` MUST expose `data-tipo-nombre`/`data-fecha-reporte` so
    `paso-offline.js` can persist them into the Dexie draft row without a
    network fetch (spec `sincronizacion-pendientes`, requirement 'Draft
    Write Captures Display Metadata')."""
    client, reporte = sesion_de_creador

    response = client.get(
        reverse("reportes_paso", args=[reporte.id, "datos-generales"])
    )
    contenido = response.content.decode()

    fecha_esperada = formats.date_format(reporte.fecha_creacion, "DATETIME_FORMAT")

    assert response.status_code == 200
    assert f'data-tipo-nombre="{reporte.tipo.nombre}"' in contenido
    assert f'data-fecha-reporte="{fecha_esperada}"' in contenido


# ---------------------------------------------------------------------------
# revision (S-09 review screen; spec `validacion-reporte`)
# ---------------------------------------------------------------------------


@pytest.fixture
def reporte_con_validaciones_factory(
    usuario_factory, tipo_con_definicion_activa_factory, reporte_factory
):
    """Build a `Reporte` whose `estructura` is `estructura_con_validaciones`
    (spec scenarios 1-6), owned by a fresh creador. Returns `(client,
    reporte)` with the creador already logged in — mirrors
    `sesion_de_creador` but lets each `revision` test choose its own
    persisted `ValorDeReporte` rows."""

    def _crear(client, estructura_con_validaciones):
        tipo, definicion = tipo_con_definicion_activa_factory(
            estructura=estructura_con_validaciones()
        )
        creador = usuario_factory(username="creador_de_revision")
        reporte = reporte_factory(tipo=tipo, definicion=definicion, creador=creador)
        client.force_login(creador)
        return reporte

    return _crear


@pytest.mark.django_db
def test_get_revision_como_creador_lista_errores_y_advertencias(
    client, estructura_con_validaciones, reporte_con_validaciones_factory, usuario_factory
):
    reporte = reporte_con_validaciones_factory(client, estructura_con_validaciones)
    autor = usuario_factory(username="autor-de-valores-revision")
    ValorDeReporte.objects.create(
        reporte=reporte,
        identificador_de_campo="observaciones-generales",
        valor="Todo en orden.",
        autor=autor,
    )
    ValorDeReporte.objects.create(
        reporte=reporte,
        identificador_de_campo="estado-general",
        valor="No cumple",
        autor=autor,
    )
    ValorDeReporte.objects.create(
        reporte=reporte, identificador_de_campo="p-01_inicio", valor="09:00", autor=autor
    )
    ValorDeReporte.objects.create(
        reporte=reporte, identificador_de_campo="p-01_fin", valor="08:00", autor=autor
    )
    # "estado-general" obligatorio filled, but as "No cumple" with no
    # observación (advertencia) and a stray hora range (advertencia). No
    # errores expected here since every obligatorio is present.

    response = client.get(reverse("reportes_revision", args=[reporte.id]))

    assert response.status_code == 200
    resultado = response.context["resultado"]
    assert resultado.errores == ()
    assert len(resultado.advertencias) == 2


@pytest.mark.django_db
def test_get_revision_reporte_de_otro_usuario_da_404(
    cliente_autenticado, usuario_factory, reporte_factory
):
    otro_creador = usuario_factory(username="otro_usuario_revision")
    reporte = reporte_factory(creador=otro_creador)

    response = cliente_autenticado.get(reverse("reportes_revision", args=[reporte.id]))

    assert response.status_code == 404


@pytest.mark.django_db
def test_get_revision_participante_invitado_accede(sesion_de_invitado):
    """Spec `cierre-reporte` — "Invited participant views revision": an
    invited non-creator user gets 200 (backlog #8, tasks.md 3.3)."""
    client, reporte = sesion_de_invitado

    response = client.get(reverse("reportes_revision", args=[reporte.id]))

    assert response.status_code == 200


@pytest.mark.django_db
def test_get_revision_no_invitado_da_404(
    cliente_autenticado, usuario_factory, reporte_factory
):
    """Spec `cierre-reporte` — "Non-invited user is denied revision access":
    a user who is neither creator nor participant gets 404 (backlog #8,
    tasks.md 3.4)."""
    otro_creador = usuario_factory(username="creador_revision_no_invitado")
    reporte = reporte_factory(creador=otro_creador)

    response = cliente_autenticado.get(reverse("reportes_revision", args=[reporte.id]))

    assert response.status_code == 404


@pytest.mark.django_db
def test_get_revision_anonimo_redirige_a_login(client, reporte_factory):
    reporte = reporte_factory()

    response = client.get(reverse("reportes_revision", args=[reporte.id]))

    assert response.status_code == 302
    assert reverse("login") in response.url


@pytest.mark.django_db
def test_get_revision_con_errores_deshabilita_generar(
    client, estructura_con_validaciones, reporte_con_validaciones_factory
):
    # No `ValorDeReporte` rows persisted at all — every obligatorio is
    # missing, so `errores` is non-empty (spec: "Errores present disables
    # Generar").
    reporte = reporte_con_validaciones_factory(client, estructura_con_validaciones)

    response = client.get(reverse("reportes_revision", args=[reporte.id]))

    assert response.status_code == 200
    assert response.context["resultado"].errores != ()
    # `disabled` moved with the closure control — change
    # `cierre-en-participantes` (tasks.md 7.1) — asserted on
    # `reportes_participantes` in `test_get_participantes_creador_ineligible_ve_disabled_y_razon`.


@pytest.mark.django_db
def test_get_revision_sin_errores_habilita_generar(
    client, estructura_con_validaciones, reporte_con_validaciones_factory, usuario_factory
):
    reporte = reporte_con_validaciones_factory(client, estructura_con_validaciones)
    autor = usuario_factory(username="autor-sin-errores")
    ValorDeReporte.objects.create(
        reporte=reporte,
        identificador_de_campo="observaciones-generales",
        valor="Todo en orden.",
        autor=autor,
    )
    ValorDeReporte.objects.create(
        reporte=reporte, identificador_de_campo="estado-general", valor="Cumple", autor=autor
    )
    ValorDeReporte.objects.create(
        reporte=reporte, identificador_de_campo="p-01_inicio", valor="08:00", autor=autor
    )
    ValorDeReporte.objects.create(
        reporte=reporte, identificador_de_campo="p-01_fin", valor="09:00", autor=autor
    )

    response = client.get(reverse("reportes_revision", args=[reporte.id]))

    assert response.status_code == 200
    assert response.context["resultado"].errores == ()
    # `disabled` moved with the closure control — change
    # `cierre-en-participantes` (tasks.md 7.2) — asserted on
    # `reportes_participantes` in `test_get_participantes_creador_elegible_sin_disabled`.


# ---------------------------------------------------------------------------
# revision — visual retrofit (change `retrofit-visual-design2` PR3, tasks.md
# Phase 4; design D3/D6, spec `visual-design-system`)
# ---------------------------------------------------------------------------


# `test_revision_boton_primario_deshabilitado_muestra_razon_via_acciones_razon`
# (originally `retrofit-visual-design2` PR3, design D6, task 4.1) folded into
# `test_get_participantes_creador_ineligible_ve_disabled_y_razon` — change
# `cierre-en-participantes` (tasks.md 7.3): the closure control, its
# `disabled` attribute, and its `.acciones__razon` sibling all moved to
# `participantes.html`.


@pytest.mark.django_db
def test_revision_tripwire_disabled_sigue_presente_como_antes(
    client,
    estructura_con_validaciones,
    tipo_con_definicion_activa_factory,
    reporte_factory,
    usuario_factory,
):
    """The literal `disabled` HTML-attribute tripwire (originally
    `retrofit-visual-design2` PR3, design D6, task 4.2) stays on
    `reportes_revision` — the "Marcar como terminado" closure control never
    moved to `participantes.html`. Builds its own two `TipoDeReporte` rows
    (distinct `codigo`) directly, since `reporte_con_validaciones_factory`
    hardcodes a single codigo/username pair and cannot be called twice in
    one test."""
    tipo_con_errores, definicion_con_errores = tipo_con_definicion_activa_factory(
        estructura=estructura_con_validaciones(),
        codigo="instalacion-resinas-tripwire-con-errores",
    )
    creador_con_errores = usuario_factory(username="creador-tripwire-con-errores")
    reporte_con_errores = reporte_factory(
        tipo=tipo_con_errores,
        definicion=definicion_con_errores,
        creador=creador_con_errores,
    )
    client.force_login(creador_con_errores)
    respuesta_con_errores = client.get(
        reverse("reportes_revision", args=[reporte_con_errores.id])
    )
    assert "disabled" in respuesta_con_errores.content.decode()

    tipo_sin_errores, definicion_sin_errores = tipo_con_definicion_activa_factory(
        estructura=estructura_con_validaciones(),
        codigo="instalacion-resinas-tripwire-sin-errores",
    )
    creador_sin_errores = usuario_factory(username="creador-tripwire-sin-errores")
    reporte_sin_errores = reporte_factory(
        tipo=tipo_sin_errores,
        definicion=definicion_sin_errores,
        creador=creador_sin_errores,
    )
    for identificador, valor in (
        ("observaciones-generales", "Todo en orden."),
        ("estado-general", "Cumple"),
        ("p-01_inicio", "08:00"),
        ("p-01_fin", "09:00"),
    ):
        ValorDeReporte.objects.create(
            reporte=reporte_sin_errores,
            identificador_de_campo=identificador,
            valor=valor,
            autor=creador_sin_errores,
        )
    client.force_login(creador_sin_errores)
    respuesta_sin_errores = client.get(
        reverse("reportes_revision", args=[reporte_sin_errores.id])
    )
    assert "disabled" not in respuesta_sin_errores.content.decode()


@pytest.mark.django_db
def test_revision_aplica_hoja_modal_component_class(
    client, estructura_con_validaciones, reporte_con_validaciones_factory
):
    """Change `retrofit-visual-design2` PR3 (design D3/D5, task 4.3): the
    errores/advertencias listing is wrapped in the `.hoja` component class
    (DESIGN2 §4 "Hoja modal (S-09)")."""
    reporte = reporte_con_validaciones_factory(client, estructura_con_validaciones)

    response = client.get(reverse("reportes_revision", args=[reporte.id]))
    contenido = response.content.decode()

    assert response.status_code == 200
    assert 'class="hoja"' in contenido
    assert "hoja__encabezado" in contenido
    assert "hoja__cuerpo" in contenido


@pytest.mark.django_db
def test_participantes_aplica_checklist_y_campo_component_classes(
    sesion_de_creador, participacion_factory
):
    """Change `retrofit-visual-design2` PR3 (design D3, task 4.4): the
    invitados listing gets the `.checklist` component class and the invite
    form's field gets `.campo` (DESIGN2 §4 "Checklist por rol" reused as the
    rhythm/list wrapper, matching PR2's `paso.html` precedent; S-10)."""
    client, reporte = sesion_de_creador
    participacion_factory(reporte, username="participante-checklist")

    response = client.get(reverse("reportes_participantes", args=[reporte.id]))
    contenido = response.content.decode()

    assert response.status_code == 200
    assert 'class="checklist' in contenido
    assert 'class="campo' in contenido


def test_sw_js_contiene_cache_v7(client):
    """Change `vista-sincronizacion-pendientes` (design D7, task 6.1/6.2):
    the SW `CACHE` version bumps `v6` -> `v7` so a returning user's stale
    cache never wins over the S-15 screen's new navigation branch. One
    assertion per PR (Testing Strategy) — supersedes the PR3 `v6` pin
    (mirrors `test_estatico.py::test_sw_js_cachea_sincronizacion_y_bumpea_v7`,
    which also asserts the new navigation-branch path)."""
    response = client.get("/sw.js")
    contenido = response.content.decode()

    assert "reportes-offline-v23" in contenido


@pytest.mark.django_db
def test_post_paso_sin_valor_obligatorio_no_bloquea(sesion_de_creador):
    """Non-blocking obligatorio marker: submitting without a value for a
    field marked `obligatorio` must NOT return a validation error, and the
    other submitted values must still be persisted (spec scenario "Missing
    obligatorio field does not block persistence")."""
    client, reporte = sesion_de_creador

    response = client.post(
        reverse("reportes_paso", args=[reporte.id, "datos-generales"]),
        data={"turno": ""},
    )

    assert response.status_code == 302
    assert not ValorDeReporte.objects.filter(
        reporte=reporte, identificador_de_campo="turno"
    ).exists()


@pytest.mark.django_db
def test_post_paso_con_rango_invalido_no_bloquea(sesion_de_creador):
    """Server-side non-blocking hora range re-check (spec: "Direct POST
    with invalid hora range still persists"): a stray `fin <= inicio` POST
    must still upsert both values and must NOT be blocked server-side —
    the design decision routes the actual check through `validar_reporte`
    at the S-09 gate, not through `paso`'s POST handler, so this test
    proves the existing non-blocking POST contract already covers it
    without any new server-side branch."""
    client, reporte = sesion_de_creador

    response = client.post(
        reverse("reportes_paso", args=[reporte.id, "proceso-instalacion"]),
        data={"p-01_inicio": "10:00", "p-01_fin": "08:00"},
    )

    assert response.status_code == 302
    inicio = ValorDeReporte.objects.get(
        reporte=reporte, identificador_de_campo="p-01_inicio"
    )
    fin = ValorDeReporte.objects.get(
        reporte=reporte, identificador_de_campo="p-01_fin"
    )
    assert inicio.valor == "10:00"
    assert fin.valor == "08:00"


# ---------------------------------------------------------------------------
# cerrar_reporte (backlog #7, task 4; spec `cierre-reporte`; design D2, D9)
# ---------------------------------------------------------------------------


@pytest.mark.django_db
def test_cerrar_reporte_no_creador_devuelve_404(
    cliente_autenticado, usuario_factory, reporte_factory
):
    otro_creador = usuario_factory(username="otro_creador_cierre")
    reporte = reporte_factory(creador=otro_creador)

    response = cliente_autenticado.post(
        reverse("reportes_cerrar", args=[reporte.id])
    )

    assert response.status_code == 404
    assert not VistoBueno.objects.filter(reporte=reporte).exists()


@pytest.mark.django_db
def test_cerrar_reporte_participante_invitado_devuelve_404(sesion_de_invitado):
    """Spec `cierre-reporte` — "Invited non-creator participant cannot
    close": `cerrar_reporte` stays creator-only, unaffected by
    `ParticipacionEnReporte` (backlog #8, tasks.md 3.9)."""
    client, reporte = sesion_de_invitado

    response = client.post(reverse("reportes_cerrar", args=[reporte.id]))

    assert response.status_code == 404
    assert not VistoBueno.objects.filter(reporte=reporte).exists()


@pytest.mark.django_db
def test_cerrar_reporte_rechazado_si_no_puede_generar(
    client, estructura_con_validaciones, reporte_con_validaciones_factory
):
    # No `ValorDeReporte` rows persisted — every obligatorio is missing, so
    # `puede_generar` is False.
    reporte = reporte_con_validaciones_factory(client, estructura_con_validaciones)

    response = client.post(reverse("reportes_cerrar", args=[reporte.id]))

    assert not VistoBueno.objects.filter(reporte=reporte).exists()
    reporte.refresh_from_db()
    assert reporte.estado == EstadoDeReporte.EN_PROGRESO
    assert response.status_code == 302
    assert response.url == reverse("reportes_revision", args=[reporte.id])


@pytest.mark.django_db
def test_cerrar_reporte_creador_exitoso(reporte_listo_para_cerrar):
    client, reporte = reporte_listo_para_cerrar

    response = client.post(reverse("reportes_cerrar", args=[reporte.id]))

    assert VistoBueno.objects.filter(
        reporte=reporte, usuario=reporte.creador
    ).exists()
    reporte.refresh_from_db()
    assert reporte.estado == EstadoDeReporte.TERMINADO
    assert response.status_code == 302
    assert response.url == reverse("reportes_mis")


@pytest.mark.django_db
def test_cerrar_reporte_doble_post_es_idempotente(reporte_listo_para_cerrar):
    client, reporte = reporte_listo_para_cerrar

    primera = client.post(reverse("reportes_cerrar", args=[reporte.id]))
    segunda = client.post(reverse("reportes_cerrar", args=[reporte.id]))

    assert primera.status_code == 302
    assert segunda.status_code == 302
    assert VistoBueno.objects.filter(reporte=reporte).count() == 1
    mensajes_segunda = list(get_messages(segunda.wsgi_request))
    assert any(mensaje.level_tag == "success" for mensaje in mensajes_segunda)


# ---------------------------------------------------------------------------
# eliminar_reporte (soft delete: creator-only, regardless of estado)
# ---------------------------------------------------------------------------


@pytest.mark.django_db
def test_eliminar_reporte_no_creador_devuelve_404(
    cliente_autenticado, usuario_factory, reporte_factory
):
    otro_creador = usuario_factory(username="otro_creador_eliminar")
    reporte = reporte_factory(creador=otro_creador)

    response = cliente_autenticado.post(
        reverse("reportes_eliminar", args=[reporte.id])
    )

    assert response.status_code == 404
    reporte.refresh_from_db()
    assert reporte.eliminado_en is None


@pytest.mark.django_db
def test_eliminar_reporte_participante_invitado_devuelve_404(sesion_de_invitado):
    """Only the creator may delete, unaffected by `ParticipacionEnReporte`
    (mirrors `cerrar_reporte`'s "Invited non-creator participant cannot
    close")."""
    client, reporte = sesion_de_invitado

    response = client.post(reverse("reportes_eliminar", args=[reporte.id]))

    assert response.status_code == 404
    reporte.refresh_from_db()
    assert reporte.eliminado_en is None


@pytest.mark.django_db
def test_eliminar_reporte_get_muestra_confirmacion_sin_efecto(sesion_de_creador):
    client, reporte = sesion_de_creador

    response = client.get(reverse("reportes_eliminar", args=[reporte.id]))

    assert response.status_code == 200
    reporte.refresh_from_db()
    assert reporte.eliminado_en is None


@pytest.mark.django_db
def test_eliminar_reporte_creador_exitoso(sesion_de_creador):
    client, reporte = sesion_de_creador

    response = client.post(reverse("reportes_eliminar", args=[reporte.id]))

    reporte.refresh_from_db()
    assert reporte.eliminado_en is not None
    assert response.status_code == 302
    assert response.url == reverse("reportes_mis")


@pytest.mark.django_db
def test_eliminar_reporte_no_borra_datos_relacionados(sesion_de_creador):
    """Soft delete only stamps `eliminado_en` — `ValorDeReporte` and every
    other related row stay intact for audit/recovery."""
    client, reporte = sesion_de_creador
    ValorDeReporte.objects.create(
        reporte=reporte,
        identificador_de_campo="turno",
        valor="Día",
        autor=reporte.creador,
    )

    client.post(reverse("reportes_eliminar", args=[reporte.id]))

    assert ValorDeReporte.objects.filter(reporte=reporte).count() == 1
    assert Reporte.objects.filter(pk=reporte.pk).exists()


@pytest.mark.django_db
def test_eliminar_reporte_terminado_tambien_se_puede_eliminar(
    reporte_listo_para_cerrar,
):
    """Deletion is allowed regardless of `estado` — `en_progreso` or
    `terminado`."""
    client, reporte = reporte_listo_para_cerrar
    client.post(reverse("reportes_cerrar", args=[reporte.id]))
    reporte.refresh_from_db()
    assert reporte.estado == EstadoDeReporte.TERMINADO

    response = client.post(reverse("reportes_eliminar", args=[reporte.id]))

    reporte.refresh_from_db()
    assert reporte.eliminado_en is not None
    assert response.status_code == 302


@pytest.mark.django_db
def test_reporte_eliminado_404_en_paso_revision_participantes(sesion_de_creador):
    """Once soft-deleted, a `Reporte` 404s exactly like one that never
    existed on every access-scoped screen — creator included."""
    client, reporte = sesion_de_creador
    client.post(reverse("reportes_eliminar", args=[reporte.id]))

    for nombre, args in (
        ("reportes_paso", [reporte.id, "datos-generales"]),
        ("reportes_revision", [reporte.id]),
        ("reportes_participantes", [reporte.id]),
    ):
        response = client.get(reverse(nombre, args=args))
        assert response.status_code == 404


@pytest.mark.django_db
def test_reporte_eliminado_no_aparece_en_mis_reportes(sesion_de_creador):
    client, reporte = sesion_de_creador
    client.post(reverse("reportes_eliminar", args=[reporte.id]))

    response = client.get(reverse("reportes_mis"))

    ids_listados = {
        tarjeta.reporte.id
        for grupo in response.context["grupos"]
        for tarjeta in grupo["tarjetas"]
    }
    assert reporte.id not in ids_listados


# ---------------------------------------------------------------------------
# generar (backlog #7, task 5; spec `generacion-documento`; design D3, D6)
# ---------------------------------------------------------------------------


@pytest.mark.django_db
def test_generar_sin_visto_bueno_redirige_con_error(reporte_listo_para_cerrar):
    # `reporte_listo_para_cerrar` is eligible but never closed — no
    # `VistoBueno` row exists yet (spec: "Generation attempted before
    # closure").
    client, reporte = reporte_listo_para_cerrar

    response = client.post(reverse("reportes_generar", args=[reporte.id]))

    assert response.status_code == 302
    assert response.url == reverse("reportes_revision", args=[reporte.id])
    assert not Generacion.objects.filter(reporte=reporte).exists()
    mensajes = list(get_messages(response.wsgi_request))
    assert any(mensaje.level_tag == "error" for mensaje in mensajes)


@pytest.mark.django_db
def test_generar_rechazado_si_no_puede_generar_pese_a_visto_bueno(
    reporte_listo_para_cerrar,
):
    # `VistoBueno` exists, but a later edit made the report ineligible
    # again (spec: "Generation rejected when no longer eligible").
    client, reporte = reporte_listo_para_cerrar
    client.post(reverse("reportes_cerrar", args=[reporte.id]))
    ValorDeReporte.objects.filter(
        reporte=reporte, identificador_de_campo="observaciones-generales"
    ).delete()

    response = client.post(reverse("reportes_generar", args=[reporte.id]))

    assert response.status_code == 302
    assert response.url == reverse("reportes_revision", args=[reporte.id])
    assert not Generacion.objects.filter(reporte=reporte).exists()
    mensajes = list(get_messages(response.wsgi_request))
    assert any(mensaje.level_tag == "error" for mensaje in mensajes)


@pytest.mark.django_db
def test_generar_participante_invitado_es_exitoso(
    reporte_listo_para_cerrar, participacion_factory
):
    """Spec `generacion-documento` — "Invited participant generates
    successfully": an invited non-creator user B succeeds once the report
    is closed (backlog #8, tasks.md 3.5; replaces half of the former
    `test_generar_no_creador_tambien_puede_generar`, per design's
    identified reversal of "Any Authenticated User May Generate")."""
    client, reporte = reporte_listo_para_cerrar
    client.post(reverse("reportes_cerrar", args=[reporte.id]))
    client.logout()
    otro = participacion_factory(reporte, username="participante-generar")
    client.force_login(otro)

    response = client.post(reverse("reportes_generar", args=[reporte.id]))

    assert response.status_code == 200
    generacion = Generacion.objects.get(reporte=reporte)
    assert generacion.usuario == otro


@pytest.mark.django_db
def test_generar_no_participante_devuelve_404(
    reporte_listo_para_cerrar, usuario_factory
):
    """Spec `generacion-documento` — "Non-participant authenticated user is
    denied": a non-creator, non-invited user C is now rejected with 404
    (backlog #8, tasks.md 3.6; captures the reversal from "Any Authenticated
    User May Generate" to creator-or-invited-participant only)."""
    client, reporte = reporte_listo_para_cerrar
    client.post(reverse("reportes_cerrar", args=[reporte.id]))
    client.logout()
    otro = usuario_factory(username="usuario-no-participante-generar")
    client.force_login(otro)

    response = client.post(reverse("reportes_generar", args=[reporte.id]))

    assert response.status_code == 404
    assert not Generacion.objects.filter(reporte=reporte).exists()


@pytest.mark.django_db
def test_generar_captura_problema_de_generacion_y_redirige(reporte_listo_para_cerrar):
    client, reporte = reporte_listo_para_cerrar
    client.post(reverse("reportes_cerrar", args=[reporte.id]))

    with mock.patch(
        "reportes.views.generador.generar_reporte",
        side_effect=PlantillaIlegible("plantilla rota"),
    ):
        response = client.post(reverse("reportes_generar", args=[reporte.id]))

    assert response.status_code == 302
    assert response.url == reverse("reportes_revision", args=[reporte.id])
    assert not Generacion.objects.filter(reporte=reporte).exists()
    mensajes = list(get_messages(response.wsgi_request))
    assert any(mensaje.level_tag == "error" for mensaje in mensajes)


@pytest.mark.django_db
def test_generar_exitoso_streamea_xlsx_con_headers_correctos(
    reporte_listo_para_cerrar,
):
    from django.utils import timezone

    client, reporte = reporte_listo_para_cerrar
    client.post(reverse("reportes_cerrar", args=[reporte.id]))

    response = client.post(reverse("reportes_generar", args=[reporte.id]))

    assert response.status_code == 200
    assert response["Content-Type"] == (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    esperado = (
        f"{reporte.tipo.codigo}-{reporte.id}-{timezone.localdate():%Y%m%d}.xlsx"
    )
    assert response["Content-Disposition"] == f'attachment; filename="{esperado}"'
    libro = load_workbook(BytesIO(response.content))
    assert libro.sheetnames == ["REPORTE"]
    assert libro["REPORTE"]["M10"].value == "Todo en orden."
    assert libro["REPORTE"]["M25"].value == "08:00"


@pytest.mark.django_db
def test_generar_repetido_crea_multiples_filas_generacion(reporte_listo_para_cerrar):
    client, reporte = reporte_listo_para_cerrar
    client.post(reverse("reportes_cerrar", args=[reporte.id]))

    primera = client.post(reverse("reportes_generar", args=[reporte.id]))
    segunda = client.post(reverse("reportes_generar", args=[reporte.id]))

    assert primera.status_code == 200
    assert segunda.status_code == 200
    assert Generacion.objects.filter(reporte=reporte).count() == 2


# ---------------------------------------------------------------------------
# Template & messages wiring (backlog #7, task 6; design D4)
# ---------------------------------------------------------------------------


@pytest.mark.django_db
def test_get_revision_con_visto_bueno_muestra_form_generar(reporte_listo_para_cerrar):
    client, reporte = reporte_listo_para_cerrar
    client.post(reverse("reportes_cerrar", args=[reporte.id]))

    response = client.get(reverse("reportes_revision", args=[reporte.id]))
    contenido = response.content.decode()

    assert response.status_code == 200
    assert f'action="{reverse("reportes_generar", args=[reporte.id])}"' in contenido
    assert "csrfmiddlewaretoken" in contenido
    assert "<form" in contenido


@pytest.mark.django_db
def test_get_revision_no_creador_no_ve_boton_cerrar(
    reporte_listo_para_cerrar, usuario_factory, rf
):
    _client, reporte = reporte_listo_para_cerrar
    otro = usuario_factory(username="otro-no-creador-revision")
    request = rf.get(reverse("reportes_revision", args=[reporte.id]))
    request.user = otro

    resultado = validar_reporte(reporte)
    contenido = render_to_string(
        "reportes/revision.html",
        {"reporte": reporte, "resultado": resultado, "tiene_visto_bueno": False},
        request=request,
    )

    assert "Marcar como terminado" not in contenido


@pytest.mark.django_db
def test_get_revision_creador_ve_form_cerrar_pero_get_no_crea_visto_bueno(
    reporte_listo_para_cerrar,
):
    """Follow-up to `cierre-en-participantes`: the closure control is now
    also reachable directly from `reportes_revision` (not only via
    `participantes.html`), so a report is fully actionable from wherever
    the creator enters it — a GET still never creates a `VistoBueno`
    (only `cerrar_reporte`'s POST does)."""
    client, reporte = reporte_listo_para_cerrar

    response = client.get(reverse("reportes_revision", args=[reporte.id]))
    contenido = response.content.decode()

    assert response.status_code == 200
    assert f'action="{reverse("reportes_cerrar", args=[reporte.id])}"' in contenido
    assert "Marcar como terminado" in contenido
    assert not VistoBueno.objects.filter(reporte=reporte).exists()


@pytest.mark.django_db
def test_get_revision_contiene_link_a_participantes_e_historial(
    reporte_listo_para_cerrar,
):
    """`reportes_revision` links to `reportes_participantes` with the
    "Participantes e historial →" label, for viewing invited users and
    change history (separate from the closure action, which now lives on
    both screens)."""
    client, reporte = reporte_listo_para_cerrar

    response = client.get(reverse("reportes_revision", args=[reporte.id]))
    contenido = response.content.decode()

    assert response.status_code == 200
    assert "Participantes e historial" in contenido
    assert (
        f'href="{reverse("reportes_participantes", args=[reporte.id])}"' in contenido
    )


@pytest.mark.django_db
def test_edicion_post_cierre_sigue_funcionando(reporte_listo_para_cerrar):
    # Spec `cierre-reporte`: "Editing a value after closure succeeds" —
    # `paso` stays fully open even once `estado == TERMINADO`.
    client, reporte = reporte_listo_para_cerrar
    client.post(reverse("reportes_cerrar", args=[reporte.id]))
    reporte.refresh_from_db()
    assert reporte.estado == EstadoDeReporte.TERMINADO

    response = client.post(
        reverse("reportes_paso", args=[reporte.id, "datos-generales"]),
        data={
            "observaciones-generales": "Actualizado tras cierre.",
            "estado-general": "Cumple",
        },
    )

    assert response.status_code == 302
    valor = ValorDeReporte.objects.get(
        reporte=reporte, identificador_de_campo="observaciones-generales"
    )
    assert valor.valor == "Actualizado tras cierre."


# ---------------------------------------------------------------------------
# invitar (backlog #8, task 4; spec `colaboracion-reporte`; design's
# "Invite view shape")
# ---------------------------------------------------------------------------


@pytest.mark.django_db
def test_invitar_exitoso(sesion_de_creador, usuario_factory):
    """Spec `colaboracion-reporte` — "Successful invite": the creator
    invites an existing, not-yet-invited user by exact username; a
    `ParticipacionEnReporte` row is created and a success flash message is
    shown (tasks.md 4.1)."""
    client, reporte = sesion_de_creador
    invitado = usuario_factory(username="invitado-exitoso")

    response = client.post(
        reverse("reportes_invitar", args=[reporte.id]),
        data={"username": invitado.username},
    )

    assert response.status_code == 302
    assert response.url == reverse("reportes_participantes", args=[reporte.id])
    assert ParticipacionEnReporte.objects.filter(
        reporte=reporte, usuario=invitado
    ).exists()
    mensajes = list(get_messages(response.wsgi_request))
    assert any(mensaje.level_tag == "success" for mensaje in mensajes)


@pytest.mark.django_db
def test_invitar_con_next_mis_redirige_a_mis_reportes(client, usuario_factory, reporte_factory):
    """The "compartir" quick-share icon on `mis_reportes.html` posts
    `next=mis` so the creator never has to open the report to invite
    someone — the response goes straight back to "Mis reportes" instead
    of `reportes_participantes`."""
    creador = usuario_factory(username="creador-next-mis")
    reporte = reporte_factory(creador=creador)
    invitado = usuario_factory(username="invitado-next-mis")
    client.force_login(creador)

    response = client.post(
        reverse("reportes_invitar", args=[reporte.id]),
        data={"username": invitado.username, "next": "mis"},
    )

    assert response.status_code == 302
    assert response.url == reverse("reportes_mis")
    assert ParticipacionEnReporte.objects.filter(
        reporte=reporte, usuario=invitado
    ).exists()


@pytest.mark.django_db
def test_invitar_idempotente(sesion_de_creador, participacion_factory):
    """Spec `colaboracion-reporte` — "Inviting an already-invited user is
    idempotent": no error, exactly one row exists after the repeat invite
    (tasks.md 4.2)."""
    client, reporte = sesion_de_creador
    invitado = participacion_factory(reporte, username="ya-invitado")

    response = client.post(
        reverse("reportes_invitar", args=[reporte.id]),
        data={"username": invitado.username},
    )

    assert response.status_code == 302
    assert (
        ParticipacionEnReporte.objects.filter(
            reporte=reporte, usuario=invitado
        ).count()
        == 1
    )


@pytest.mark.django_db
def test_invitar_usuario_inexistente(sesion_de_creador):
    """Spec `colaboracion-reporte` — "Inviting a nonexistent username": no
    row is created and an error flash message is shown (tasks.md 4.3)."""
    client, reporte = sesion_de_creador

    response = client.post(
        reverse("reportes_invitar", args=[reporte.id]),
        data={"username": "nadie"},
    )

    assert response.status_code == 302
    assert not ParticipacionEnReporte.objects.filter(reporte=reporte).exists()
    mensajes = list(get_messages(response.wsgi_request))
    assert any(mensaje.level_tag == "error" for mensaje in mensajes)


@pytest.mark.django_db
def test_invitar_no_creador_devuelve_404(
    cliente_autenticado, reporte_factory, usuario_factory
):
    """Spec `colaboracion-reporte` — "Non-creator cannot invite": a
    non-creator, non-participant authenticated user gets 404 and no row is
    created (tasks.md 4.4)."""
    otro_creador = usuario_factory(username="creador-invitacion-rechazada")
    reporte = reporte_factory(creador=otro_creador)
    objetivo = usuario_factory(username="objetivo-invitacion-rechazada")

    response = cliente_autenticado.post(
        reverse("reportes_invitar", args=[reporte.id]),
        data={"username": objetivo.username},
    )

    assert response.status_code == 404
    assert not ParticipacionEnReporte.objects.filter(reporte=reporte).exists()


@pytest.mark.django_db
def test_invitar_a_si_mismo_rechazado(sesion_de_creador):
    """Design's self-invite rejection: protects "creator has no
    participation row" — no row is created for the creator, error flash
    message shown (tasks.md 4.5)."""
    client, reporte = sesion_de_creador

    response = client.post(
        reverse("reportes_invitar", args=[reporte.id]),
        data={"username": reporte.creador.username},
    )

    assert response.status_code == 302
    assert not ParticipacionEnReporte.objects.filter(
        reporte=reporte, usuario=reporte.creador
    ).exists()
    mensajes = list(get_messages(response.wsgi_request))
    assert any(mensaje.level_tag == "error" for mensaje in mensajes)


# ---------------------------------------------------------------------------
# participantes (backlog #8, task 4; spec `colaboracion-reporte` —
# "Participants and History View")
# ---------------------------------------------------------------------------


@pytest.mark.django_db
def test_participantes_lista_invitados_y_creador(
    client, reporte_con_participantes_factory
):
    """Spec `colaboracion-reporte` — "View lists participants and creator
    label": the invited user's username is listed, the creator is shown
    labeled as creator (tasks.md 4.6)."""
    reporte, invitados = reporte_con_participantes_factory(n=1)
    client.force_login(reporte.creador)

    response = client.get(reverse("reportes_participantes", args=[reporte.id]))
    contenido = response.content.decode()

    assert response.status_code == 200
    assert invitados[0].username in contenido
    assert reporte.creador.username in contenido


@pytest.mark.django_db
def test_participantes_historial_mas_reciente_primero(sesion_de_creador):
    """Spec `colaboracion-reporte` — "History renders most-recent-first"
    (tasks.md 4.7)."""
    client, reporte = sesion_de_creador
    primero = CambioDeValor.objects.create(
        reporte=reporte,
        identificador_de_campo="campo-1",
        valor_anterior=None,
        autor=reporte.creador,
    )
    segundo = CambioDeValor.objects.create(
        reporte=reporte,
        identificador_de_campo="campo-2",
        valor_anterior=None,
        autor=reporte.creador,
    )

    response = client.get(reverse("reportes_participantes", args=[reporte.id]))

    sesiones = response.context["sesiones_de_cambios"]
    cambios = [item["cambio"] for sesion in sesiones for item in sesion["items"]]
    assert cambios == [segundo, primero]


@pytest.mark.django_db
def test_participantes_no_participante_devuelve_404(
    cliente_autenticado, reporte_factory, usuario_factory
):
    """Non-creator, non-participant user gets 404 on the participantes view
    (tasks.md 4.8)."""
    otro_creador = usuario_factory(username="creador-participantes-rechazado")
    reporte = reporte_factory(creador=otro_creador)

    response = cliente_autenticado.get(
        reverse("reportes_participantes", args=[reporte.id])
    )

    assert response.status_code == 404


@pytest.mark.django_db
def test_participantes_contexto_incluye_resultado(sesion_de_creador):
    """Change `cierre-en-participantes` (tasks.md 1.3; design's Interfaces
    section): `participantes` computes `resultado` via `validar_reporte`,
    exactly like `revision` already does, so the closure form can render
    the `puede_generar`/errores-driven gating."""
    client, reporte = sesion_de_creador

    response = client.get(reverse("reportes_participantes", args=[reporte.id]))

    assert response.status_code == 200
    assert response.context["resultado"] == validar_reporte(reporte)


# ---------------------------------------------------------------------------
# reportes_cerrar — closure endpoint (no longer exposed via a form on
# `reportes_participantes`; the closure control lives only on
# `reportes_revision` now). The endpoint itself is unchanged.
# ---------------------------------------------------------------------------


@pytest.mark.django_db
def test_post_reportes_cerrar_creador_exitoso(reporte_listo_para_cerrar):
    """Mirrors `test_cerrar_reporte_creador_exitoso`: POSTing to
    `reportes_cerrar` creates the `VistoBueno`, sets `estado=TERMINADO`,
    and redirects to `reportes_mis`."""
    client, reporte = reporte_listo_para_cerrar

    response = client.post(reverse("reportes_cerrar", args=[reporte.id]))

    assert VistoBueno.objects.filter(
        reporte=reporte, usuario=reporte.creador
    ).exists()
    reporte.refresh_from_db()
    assert reporte.estado == EstadoDeReporte.TERMINADO
    assert response.status_code == 302
    assert response.url == reverse("reportes_mis")


# ---------------------------------------------------------------------------
# service_worker (change `capa-offline`; spec `capa-offline` — "Root-Scoped
# Service Worker Route"; design's Decision "sw.js served as a Django
# template"). TDD-covered per tasks.md Phase 1 — the only client-adjacent
# surface with automated coverage in this change.
# ---------------------------------------------------------------------------


@pytest.mark.django_db
def test_sw_js_headers_correctos(client):
    """Spec: "/sw.js is served with correct headers" — 200, correct
    Content-Type, and the `Service-Worker-Allowed: /` header that grants
    root scope from outside WhiteNoise's `/static/` prefix."""
    response = client.get("/sw.js")

    assert response.status_code == 200
    assert response["Content-Type"] in (
        "application/javascript",
        "text/javascript",
    )
    assert response["Service-Worker-Allowed"] == "/"


@pytest.mark.django_db
def test_sw_js_anonimo_no_redirige_a_login(client):
    """Spec: "/sw.js is reachable without authentication" — registration
    happens before any session context is guaranteed, so this route must
    never redirect to login nor 401/403 for an anonymous client."""
    response = client.get("/sw.js")

    assert response.status_code == 200


@pytest.mark.django_db
def test_sw_js_body_referencia_paso_js(client):
    """Design's Decision "sw.js served as a Django template": the body must
    be rendered via `{% static %}`, not read from a static file — this
    proves the template actually renders rather than being served as a raw
    file, by asserting the resolved `paso.js` static URL appears in body."""
    response = client.get("/sw.js")
    contenido = response.content.decode()

    assert "/static/reportes/paso.js" in contenido


@pytest.mark.django_db
def test_base_html_incluye_ambos_links_css(client, usuario_factory):
    """`base.html` links both `static/css/tokens.css` and
    `static/css/components.css` (spec `visual-design-system`, scenario
    'base.html links the new stylesheets and font')."""
    usuario_factory(username="usuario-base-html-links")
    response = client.get(reverse("login"))
    contenido = response.content.decode()

    assert '<link rel="stylesheet" href="/static/css/tokens.css">' in contenido
    assert (
        '<link rel="stylesheet" href="/static/css/components.css">' in contenido
    )


@pytest.mark.django_db
def test_base_incluye_script_conexion_chip_defer(client, usuario_factory):
    """Change `chip-conexion-en-vivo` (tasks.md 2.2; design's Decision
    "Script included once in base.html, not per template"): the rendered
    `<head>` must include `js/conexion-chip.js` with `defer`."""
    usuario_factory(username="usuario-base-html-conexion-chip")
    response = client.get(reverse("login"))
    contenido = response.content.decode()

    assert "js/conexion-chip.js" in contenido
    assert "defer" in contenido


def test_base_html_no_referencia_cdn(client):
    """No `<link>`/`<script>` in the rendered page references a third-party
    CDN font or stylesheet (spec: 'no third-party CDN dependency')."""
    response = client.get(reverse("login"))
    contenido = response.content.decode().lower()

    assert "fonts.googleapis" not in contenido
    assert "fonts.gstatic" not in contenido
    assert "cdn." not in contenido


def test_login_primer_boton_submit_es_el_del_formulario_de_login(client):
    """Submit-order guard against the new `base.html` shell (design: a
    header/logout form ahead of the page's own submit form would silently
    break `querySelector`-based JS elsewhere, e.g. `paso.js:63`). The first
    `form button[type="submit"]` in document order on `/login/` must be the
    login form's own submit button."""
    response = client.get(reverse("login"))
    contenido = response.content.decode()

    primer_boton = re.search(r'<button[^>]*type="submit"[^>]*>([^<]*)</button>', contenido)

    assert primer_boton is not None
    assert "Ingresar" in primer_boton.group(1)


# ---------------------------------------------------------------------------
# paso — servidor_actualizado context (change `capa-offline`; spec
# `capa-offline` — client relies on this to decide whether a local draft is
# newer than what the server already has). TDD-covered per tasks.md Phase 2.
# ---------------------------------------------------------------------------


@pytest.mark.django_db
def test_get_paso_incluye_servidor_actualizado(sesion_de_creador):
    """`GET paso` must render `data-servidor-actualizado` on the form,
    derived from `max(ValorDeReporte.fecha)` for that `(reporte_id,
    seccion_id)` — the sole signal `paso-offline.js` uses to decide whether
    a local IndexedDB draft is newer than server-rendered data."""
    client, reporte = sesion_de_creador
    valor = ValorDeReporte.objects.create(
        reporte=reporte,
        identificador_de_campo="turno",
        valor="Día",
        autor=reporte.creador,
    )

    response = client.get(reverse("reportes_paso", args=[reporte.id, "datos-generales"]))
    contenido = response.content.decode()

    assert response.status_code == 200
    assert "data-servidor-actualizado=" in contenido
    esperado = valor.fecha.isoformat()
    assert esperado in contenido


@pytest.mark.django_db
def test_post_paso_actualiza_servidor_actualizado_en_siguiente_get(sesion_de_creador):
    """After a successful POST to `paso`, the subsequent GET of the SAME
    section must show `data-servidor-actualizado` updated to the new
    `max(ValorDeReporte.fecha)` — proves the value is recomputed per
    request, not cached/stale."""
    client, reporte = sesion_de_creador

    response_antes = client.get(
        reverse("reportes_paso", args=[reporte.id, "datos-generales"])
    )
    assert 'data-servidor-actualizado=""' in response_antes.content.decode()

    client.post(
        reverse("reportes_paso", args=[reporte.id, "datos-generales"]),
        data={"turno": "Noche"},
    )

    response_despues = client.get(
        reverse("reportes_paso", args=[reporte.id, "datos-generales"])
    )
    contenido_despues = response_despues.content.decode()
    valor = ValorDeReporte.objects.get(reporte=reporte, identificador_de_campo="turno")

    assert response_despues.status_code == 200
    assert valor.fecha.isoformat() in contenido_despues


# ---------------------------------------------------------------------------
# mis_reportes (backlog #12, spec `listado-reportes`; design's Data Flow /
# D1-D5). Status-bucket grouping replaces the old creador/participante
# split (design's Technical Approach note: "the delta spec supersedes
# proposal.md"). Focused command: `pytest reportes/tests/test_views.py -q
# -k mis_reportes`.
# ---------------------------------------------------------------------------


def _bucket_de(response, reporte):
    """Test-local helper: find which `grupos` bucket id contains `reporte`,
    or `None` if it appears in no bucket on the rendered page."""
    for grupo in response.context["grupos"]:
        if any(tarjeta.reporte.pk == reporte.pk for tarjeta in grupo["tarjetas"]):
            return grupo["id"]
    return None


@pytest.mark.django_db
def test_mis_reportes_anonimo_redirige_a_login(client, reporte_factory):
    """Spec 'Anonymous user is redirected'."""
    reporte_factory()

    response = client.get(reverse("reportes_mis"))

    assert response.status_code == 302
    assert reverse("login") in response.url
    assert Reporte.objects.count() == 1


@pytest.mark.django_db
def test_mis_reportes_lista_solo_accesibles(
    client, usuario_factory, tipo_con_definicion_activa_factory, reporte_factory
):
    """Spec 'Access-Scoped Report List': creator A sees R1 (own) and R2
    (invited), not R3 (stranger's)."""
    a = usuario_factory(username="mis-reportes-a")
    b = usuario_factory(username="mis-reportes-b")
    tipo1, definicion1 = tipo_con_definicion_activa_factory(
        nombre="Accesibles 1", codigo="accesibles-1"
    )
    tipo2, definicion2 = tipo_con_definicion_activa_factory(
        nombre="Accesibles 2", codigo="accesibles-2"
    )
    tipo3, definicion3 = tipo_con_definicion_activa_factory(
        nombre="Accesibles 3", codigo="accesibles-3"
    )
    r1 = reporte_factory(creador=a, tipo=tipo1, definicion=definicion1)
    r2 = reporte_factory(creador=b, tipo=tipo2, definicion=definicion2)
    ParticipacionEnReporte.objects.create(reporte=r2, usuario=a)
    r3 = reporte_factory(creador=b, tipo=tipo3, definicion=definicion3)
    client.force_login(a)

    response = client.get(reverse("reportes_mis"))
    reportes_en_pagina = [t.reporte for t in response.context["page_obj"]]

    assert response.status_code == 200
    assert r1 in reportes_en_pagina
    assert r2 in reportes_en_pagina
    assert r3 not in reportes_en_pagina


@pytest.mark.django_db
def test_mis_reportes_admin_sin_relacion_no_ve_reporte_ajeno(
    client, usuario_factory, reporte_factory
):
    """Spec 'Admin Override Explicitly Out of Scope': a staff/admin user
    with no creador/participacion relation to R4 does not see it."""
    admin = usuario_factory(username="mis-reportes-admin", is_staff=True, is_superuser=True)
    otro = usuario_factory(username="mis-reportes-creador-r4")
    r4 = reporte_factory(creador=otro)
    client.force_login(admin)

    response = client.get(reverse("reportes_mis"))
    reportes_en_pagina = [t.reporte for t in response.context["page_obj"]]

    assert response.status_code == 200
    assert r4 not in reportes_en_pagina


@pytest.mark.django_db
def test_mis_reportes_relacion_creados_filtra_antes_de_agrupar(
    client, usuario_factory, tipo_con_definicion_activa_factory, reporte_factory
):
    """Spec 'Filter restricts before grouping': user A created R1 (missing
    fields) and was invited to R2 (visto bueno present); `?relacion=creados`
    must leave R2 out of every bucket, not just its own bucket."""
    a = usuario_factory(username="mis-reportes-relacion-a")
    tipo1, definicion1 = tipo_con_definicion_activa_factory(
        nombre="Relacion filtra 1", codigo="relacion-filtra-1"
    )
    tipo2, definicion2 = tipo_con_definicion_activa_factory(
        nombre="Relacion filtra 2", codigo="relacion-filtra-2"
    )
    r1 = reporte_factory(creador=a, tipo=tipo1, definicion=definicion1)
    r2 = reporte_factory(tipo=tipo2, definicion=definicion2)
    ParticipacionEnReporte.objects.create(reporte=r2, usuario=a)
    VistoBueno.objects.create(reporte=r2, usuario=r2.creador)
    client.force_login(a)

    response = client.get(reverse("reportes_mis"), {"relacion": "creados"})

    assert response.status_code == 200
    assert _bucket_de(response, r1) is not None
    assert _bucket_de(response, r2) is None


@pytest.mark.django_db
def test_mis_reportes_relacion_por_defecto_es_todos(
    client, usuario_factory, reporte_factory, participacion_factory,
    tipo_con_definicion_activa_factory,
):
    """Spec 'Default is todos': with no `?relacion=`, both created and
    shared reports are considered for bucketing."""
    tipo1, definicion1 = tipo_con_definicion_activa_factory(
        nombre="Relacion default 1", codigo="relacion-default-1"
    )
    tipo2, definicion2 = tipo_con_definicion_activa_factory(
        nombre="Relacion default 2", codigo="relacion-default-2"
    )
    r_creado = reporte_factory(tipo=tipo1, definicion=definicion1)
    a = participacion_factory(r_creado, username="mis-reportes-relacion-default-a")
    r_creado_por_a = reporte_factory(creador=a, tipo=tipo2, definicion=definicion2)
    client.force_login(a)

    response = client.get(reverse("reportes_mis"))

    assert response.status_code == 200
    assert _bucket_de(response, r_creado) is not None
    assert _bucket_de(response, r_creado_por_a) is not None


@pytest.mark.django_db
def test_mis_reportes_bucket_terminado_es_el_mismo_para_creador_e_invitado(
    client, usuario_factory, reporte_factory, participacion_factory
):
    """Spec 'Closed report is terminado for any viewer': grouping does not
    depend on the requesting user's own authorship history."""
    reporte = reporte_factory()
    invitado = participacion_factory(reporte, username="mis-reportes-bucket-invitado")
    VistoBueno.objects.create(reporte=reporte, usuario=reporte.creador)

    client.force_login(reporte.creador)
    respuesta_creador = client.get(reverse("reportes_mis"))
    client.force_login(invitado)
    respuesta_invitado = client.get(reverse("reportes_mis"))

    assert _bucket_de(respuesta_creador, reporte) == "terminado"
    assert _bucket_de(respuesta_invitado, reporte) == "terminado"


@pytest.mark.django_db
def test_mis_reportes_en_progreso_sin_importar_quien_completo(
    client, usuario_factory, estructura_con_validaciones, tipo_con_definicion_activa_factory
):
    """Spec 'Missing fields groups as en progreso regardless of authorship':
    the creador authored one value, invited user B never authored any, and
    the report is still 'en progreso' for B — the SAME bucket as for the
    creador."""
    tipo, definicion = tipo_con_definicion_activa_factory(
        estructura=estructura_con_validaciones(),
        nombre="En progreso autoria",
        codigo="en-progreso-autoria",
    )
    creador = usuario_factory(username="mis-reportes-en-progreso-creador")
    reporte = Reporte.objects.create(tipo=tipo, definicion=definicion, creador=creador)
    ValorDeReporte.objects.create(
        reporte=reporte,
        identificador_de_campo="estado-general",
        valor="Cumple",
        autor=creador,
    )
    b = usuario_factory(username="mis-reportes-en-progreso-b")
    ParticipacionEnReporte.objects.create(reporte=reporte, usuario=b)
    client.force_login(b)

    response = client.get(reverse("reportes_mis"))

    assert _bucket_de(response, reporte) == "en_progreso"


@pytest.mark.django_db
def test_mis_reportes_filtro_estado(
    client, usuario_factory, tipo_con_definicion_activa_factory, reporte_factory
):
    """Spec 'Filter by computed estado bucket': `?estado=terminado` narrows
    to terminado reports only."""
    a = usuario_factory(username="mis-reportes-filtro-estado")
    tipo1, definicion1 = tipo_con_definicion_activa_factory(
        nombre="Filtro estado 1", codigo="filtro-estado-1"
    )
    tipo2, definicion2 = tipo_con_definicion_activa_factory(
        nombre="Filtro estado 2", codigo="filtro-estado-2"
    )
    en_progreso = reporte_factory(creador=a, tipo=tipo1, definicion=definicion1)
    terminado = reporte_factory(creador=a, tipo=tipo2, definicion=definicion2)
    VistoBueno.objects.create(reporte=terminado, usuario=a)
    client.force_login(a)

    response = client.get(reverse("reportes_mis"), {"estado": "terminado"})
    reportes_en_pagina = [t.reporte for t in response.context["page_obj"]]

    assert response.status_code == 200
    assert reportes_en_pagina == [terminado]
    assert en_progreso not in reportes_en_pagina


@pytest.mark.django_db
def test_mis_reportes_estado_terminado_encuentra_reporte_en_pagina_2_sin_filtro(
    client, usuario_factory, tipo_con_definicion_activa_factory, reporte_factory
):
    """Design D2's rationale: bucket-then-filter-then-paginate means
    `?estado=terminado` finds a terminado report that would have landed on
    page 2 of the UNFILTERED set — proves filtering happens before, not
    after, pagination."""
    a = usuario_factory(username="mis-reportes-pagina-2-estado")
    tipo, definicion = tipo_con_definicion_activa_factory(
        nombre="Pagina 2 estado", codigo="pagina-2-estado"
    )
    for _ in range(20):
        reporte_factory(creador=a, tipo=tipo, definicion=definicion)
    terminado = reporte_factory(creador=a, tipo=tipo, definicion=definicion)
    VistoBueno.objects.create(reporte=terminado, usuario=a)
    client.force_login(a)

    response = client.get(reverse("reportes_mis"), {"estado": "terminado"})
    reportes_en_pagina = [t.reporte for t in response.context["page_obj"]]

    assert response.status_code == 200
    assert reportes_en_pagina == [terminado]


@pytest.mark.django_db
def test_mis_reportes_estado_invalido_no_falla(
    client, usuario_factory, tipo_con_definicion_activa_factory, reporte_factory
):
    """Design D3, spec 'unrecognized ?estado= MUST NOT raise an error':
    `?estado=basura` returns 200 with the full unfiltered set."""
    a = usuario_factory(username="mis-reportes-estado-invalido")
    tipo1, definicion1 = tipo_con_definicion_activa_factory(
        nombre="Estado invalido 1", codigo="estado-invalido-1"
    )
    tipo2, definicion2 = tipo_con_definicion_activa_factory(
        nombre="Estado invalido 2", codigo="estado-invalido-2"
    )
    reporte_factory(creador=a, tipo=tipo1, definicion=definicion1)
    terminado = reporte_factory(creador=a, tipo=tipo2, definicion=definicion2)
    VistoBueno.objects.create(reporte=terminado, usuario=a)
    client.force_login(a)

    response = client.get(reverse("reportes_mis"), {"estado": "basura"})

    assert response.status_code == 200
    assert len(response.context["page_obj"]) == 2


@pytest.mark.django_db
def test_mis_reportes_orden_mas_reciente_primero(
    client, usuario_factory, tipo_con_definicion_activa_factory, reporte_factory
):
    """Spec 'Most recent report appears first'."""
    from datetime import timedelta

    from django.utils import timezone

    a = usuario_factory(username="mis-reportes-orden")
    tipo1, definicion1 = tipo_con_definicion_activa_factory(
        nombre="Orden 1", codigo="orden-1"
    )
    tipo2, definicion2 = tipo_con_definicion_activa_factory(
        nombre="Orden 2", codigo="orden-2"
    )
    tipo3, definicion3 = tipo_con_definicion_activa_factory(
        nombre="Orden 3", codigo="orden-3"
    )
    r1 = reporte_factory(creador=a, tipo=tipo1, definicion=definicion1)
    r2 = reporte_factory(creador=a, tipo=tipo2, definicion=definicion2)
    r3 = reporte_factory(creador=a, tipo=tipo3, definicion=definicion3)
    ahora = timezone.now()
    Reporte.objects.filter(pk=r1.pk).update(fecha_creacion=ahora - timedelta(days=2))
    Reporte.objects.filter(pk=r2.pk).update(fecha_creacion=ahora - timedelta(days=1))
    Reporte.objects.filter(pk=r3.pk).update(fecha_creacion=ahora)
    client.force_login(a)

    response = client.get(reverse("reportes_mis"))
    reportes_en_pagina = [t.reporte for t in response.context["page_obj"]]

    assert response.status_code == 200
    assert reportes_en_pagina == [r3, r2, r1]


@pytest.mark.django_db
def test_mis_reportes_pagina_1_tiene_20_y_pagina_2_tiene_1(
    client, usuario_factory, tipo_con_definicion_activa_factory, reporte_factory
):
    """Spec 'Results beyond one page are paginated', design D2 page size =
    20."""
    a = usuario_factory(username="mis-reportes-paginacion")
    tipo, definicion = tipo_con_definicion_activa_factory(
        nombre="Paginación", codigo="paginacion-mis-reportes"
    )
    for _ in range(21):
        reporte_factory(creador=a, tipo=tipo, definicion=definicion)
    client.force_login(a)

    respuesta_pagina_1 = client.get(reverse("reportes_mis"))
    respuesta_pagina_2 = client.get(reverse("reportes_mis"), {"page": "2"})

    assert respuesta_pagina_1.status_code == 200
    assert len(respuesta_pagina_1.context["page_obj"]) == 20
    assert respuesta_pagina_2.status_code == 200
    assert len(respuesta_pagina_2.context["page_obj"]) == 1


@pytest.mark.django_db
def test_mis_reportes_page_param_invalido_no_falla(
    client, usuario_factory, reporte_factory
):
    """Design D2 `get_page` behavior: `?page=abc` and `?page=999` both
    clamp to a valid page instead of raising."""
    a = usuario_factory(username="mis-reportes-page-invalido")
    reporte_factory(creador=a)
    client.force_login(a)

    respuesta_no_numerica = client.get(reverse("reportes_mis"), {"page": "abc"})
    respuesta_fuera_de_rango = client.get(reverse("reportes_mis"), {"page": "999"})

    assert respuesta_no_numerica.status_code == 200
    assert respuesta_fuera_de_rango.status_code == 200


@pytest.mark.django_db
def test_mis_reportes_pagina_2_preserva_query_string(
    client, usuario_factory, tipo_con_definicion_activa_factory, reporte_factory
):
    """Design's `{% querystring %}` note: `?page=2&q=x` keeps `q=x` in the
    pagination links rendered in the body."""
    a = usuario_factory(username="mis-reportes-querystring")
    tipo, definicion = tipo_con_definicion_activa_factory(
        nombre="Consulta preservada", codigo="querystring-mis-reportes"
    )
    for _ in range(21):
        reporte_factory(creador=a, tipo=tipo, definicion=definicion)
    client.force_login(a)

    response = client.get(
        reverse("reportes_mis"), {"page": "2", "q": "mis-reportes-querystring"}
    )
    contenido = response.content.decode()

    assert response.status_code == 200
    assert "q=mis-reportes-querystring" in contenido


@pytest.mark.django_db
def test_mis_reportes_muestra_numero_registro_asignado(
    client, usuario_factory, reporte_factory
):
    """Spec 'Assigned numero_registro renders'."""
    a = usuario_factory(username="mis-reportes-numero-registro")
    reporte = reporte_factory(creador=a)
    client.force_login(a)

    response = client.get(reverse("reportes_mis"))
    contenido = response.content.decode()

    assert response.status_code == 200
    assert str(reporte.numero_registro) in contenido


@pytest.mark.django_db
def test_mis_reportes_local_chip_cuando_numero_registro_es_none(
    client, usuario_factory, reporte_factory
):
    """Spec 'Unsynced report renders local chip' (design D4 — a persisted
    `Reporte.numero_registro` can never itself be `None`, so this exercises
    the template contract directly by injecting a `TarjetaDeReporte` whose
    `numero_registro` is `None`, the shape a future offline row from
    `vista-sincronizacion-pendientes` would have)."""
    from reportes.listado import TarjetaDeReporte

    a = usuario_factory(username="mis-reportes-local-chip")
    reporte = reporte_factory(creador=a)
    client.force_login(a)

    tarjeta_local = TarjetaDeReporte(
        reporte=reporte, bucket="en_progreso", avance=0, numero_registro=None
    )
    with mock.patch("reportes.listado.construir_tarjetas", return_value=[tarjeta_local]):
        response = client.get(reverse("reportes_mis"))
    contenido = response.content.decode()

    assert response.status_code == 200
    assert "local" in contenido


@pytest.mark.django_db
def test_mis_reportes_muestra_porcentaje_de_avance(
    client, usuario_factory, estructura_con_validaciones, tipo_con_definicion_activa_factory
):
    """Spec 'Partial completion renders a percentage': 1 of 4 obligatorios
    filled ⇒ 25%."""
    tipo, definicion = tipo_con_definicion_activa_factory(
        estructura=estructura_con_validaciones(),
        nombre="Avance porcentaje",
        codigo="avance-porcentaje",
    )
    a = usuario_factory(username="mis-reportes-avance")
    reporte = Reporte.objects.create(tipo=tipo, definicion=definicion, creador=a)
    ValorDeReporte.objects.create(
        reporte=reporte,
        identificador_de_campo="estado-general",
        valor="Cumple",
        autor=a,
    )
    client.force_login(a)

    response = client.get(reverse("reportes_mis"))
    contenido = response.content.decode()

    assert response.status_code == 200
    assert "25" in contenido


@pytest.mark.django_db
def test_mis_reportes_cta_nuevo_reporte_presente_incluso_sin_resultados(
    client, usuario_factory
):
    """Spec 'Fixed Nuevo Reporte Entry Point' / 'CTA is always present': the
    entry point moved from an inline button into the sidebar's "Reportes"
    link (always rendered via base.html for every authenticated screen),
    to avoid duplicating it with the sidebar nav — still renders even with
    zero accessible reports."""
    a = usuario_factory(username="mis-reportes-cta-vacio")
    client.force_login(a)

    response = client.get(reverse("reportes_mis"))
    contenido = response.content.decode()

    assert response.status_code == 200
    assert reverse("reportes_seleccion_tipo") in contenido


@pytest.mark.django_db
def test_mis_reportes_cta_nuevo_reporte_presente_con_filtros_y_busqueda(
    client, usuario_factory, reporte_factory
):
    """Spec 'CTA is always present': the sidebar "Reportes" entry point
    still renders under `?q=`/`?relacion=`/`?estado=` combinations,
    including ones that empty the result set."""
    a = usuario_factory(username="mis-reportes-cta-filtros")
    reporte_factory(creador=a)
    client.force_login(a)

    response = client.get(
        reverse("reportes_mis"),
        {"q": "algo-que-no-existe", "relacion": "compartidos", "estado": "terminado"},
    )
    contenido = response.content.decode()

    assert response.status_code == 200
    assert reverse("reportes_seleccion_tipo") in contenido


@pytest.mark.django_db
def test_mis_reportes_numero_de_consultas_no_crece_con_n(
    client, usuario_factory, tipo_con_definicion_activa_factory, reporte_factory
    ):
    """Task 3.3 — the view's query count must stay constant regardless of
    how many reports are being listed (design D2: `annotate(Exists(...))` +
    `select_related`/`prefetch_related` instead of one query per report)."""
    from django.test.utils import CaptureQueriesContext
    from django.db import connection

    a = usuario_factory(username="mis-reportes-num-queries")
    tipo, definicion = tipo_con_definicion_activa_factory(
        nombre="Num queries", codigo="num-queries"
    )
    reporte_factory(creador=a, tipo=tipo, definicion=definicion)
    client.force_login(a)

    with CaptureQueriesContext(connection) as capturado_uno:
        client.get(reverse("reportes_mis"))
    consultas_con_uno = len(capturado_uno.captured_queries)

    for _ in range(9):
        reporte_factory(creador=a, tipo=tipo, definicion=definicion)

    with CaptureQueriesContext(connection) as capturado_diez:
        client.get(reverse("reportes_mis"))
    consultas_con_diez = len(capturado_diez.captured_queries)

    assert consultas_con_diez == consultas_con_uno


# ---------------------------------------------------------------------------
# Live connection chip (change `chip-conexion-en-vivo`; spec `capa-offline`
# — "Live Connection Chip in Shared Screen Bar"; design's File Changes /
# "Chip renders hidden, JS reveals it"). Focused command: `pytest
# reportes/tests/test_views.py reportes/tests/test_estatico.py`.
# ---------------------------------------------------------------------------


@pytest.mark.django_db
def test_chip_conexion_presente_en_paso_mis_reportes_adjuntos_en_orden_disenio2(
    client, usuario_factory, reporte_factory
):
    """Change `chip-conexion-en-vivo` (tasks.md 2.4; DESIGN2 §4 bar order:
    volver · título · indicador · conexión · avatar): the `[data-chip-
    conexion]` node is present on the `paso`, `mis_reportes`, and `adjuntos`
    screens, in that document order relative to the bar's other elements."""
    creador = usuario_factory(username="chip-conexion-orden")
    reporte = reporte_factory(creador=creador)
    client.force_login(creador)

    respuesta_paso = client.get(
        reverse("reportes_paso", args=[reporte.id, "datos-generales"])
    )
    contenido_paso = respuesta_paso.content.decode()
    assert respuesta_paso.status_code == 200
    assert "data-chip-conexion" in contenido_paso
    assert contenido_paso.index(
        "barra-pantalla__indicador"
    ) < contenido_paso.index("data-chip-conexion")

    respuesta_mis_reportes = client.get(reverse("reportes_mis"))
    contenido_mis_reportes = respuesta_mis_reportes.content.decode()
    assert respuesta_mis_reportes.status_code == 200
    assert "data-chip-conexion" in contenido_mis_reportes
    assert contenido_mis_reportes.index(
        "data-chip-conexion"
    ) < contenido_mis_reportes.index("barra-pantalla__avatar")

    respuesta_adjuntos = client.get(reverse("reportes_adjuntos", args=[reporte.id]))
    contenido_adjuntos = respuesta_adjuntos.content.decode()
    assert respuesta_adjuntos.status_code == 200
    assert "data-chip-conexion" in contenido_adjuntos
    assert contenido_adjuntos.index(
        "barra-pantalla__titulo"
    ) < contenido_adjuntos.index("data-chip-conexion")


def test_chip_conexion_ausente_en_login(client):
    """Change `chip-conexion-en-vivo` (tasks.md 2.5; spec scenario 'Chip
    does not appear on the login screen'): `/login/` has no
    `.barra-pantalla`, so `data-chip-conexion` must not appear there."""
    response = client.get(reverse("login"))
    contenido = response.content.decode()

    assert response.status_code == 200
    assert "data-chip-conexion" not in contenido


@pytest.mark.django_db
def test_paso_offline_banner_markup_no_afectado_por_chip(
    client, usuario_factory, reporte_factory
):
    """Change `chip-conexion-en-vivo` (tasks.md 4.1; spec scenario 'Chip is
    independent from the paso-offline draft banner'): the chip addition must
    not touch `paso.html`'s `paso-offline.js` script include or its
    `data-reporte-id`/`data-seccion-id`/`data-servidor-actualizado` contract
    that `[data-borrador-banner]`/`[data-borrador-prompt]` depend on at
    runtime."""
    creador = usuario_factory(username="chip-conexion-no-afecta-borrador")
    reporte = reporte_factory(creador=creador)
    client.force_login(creador)

    response = client.get(
        reverse("reportes_paso", args=[reporte.id, "datos-generales"])
    )
    contenido = response.content.decode()

    assert response.status_code == 200
    assert "reportes/paso-offline.js" in contenido
    assert "data-reporte-id=" in contenido
    assert "data-seccion-id=" in contenido
    assert "data-servidor-actualizado=" in contenido


@pytest.mark.django_db
def test_mis_reportes_aplica_barra_pantalla_y_lista_component_classes(
    client, usuario_factory, reporte_factory
):
    """Change `retrofit-visual-design2` PR2 (design D3, task 3.9): S-02's
    screen header uses `.barra-pantalla` (DESIGN2 §4 "Barra de pantalla",
    no volver — root screen) and its report groupings use `.lista`, a
    mobile card list (DESIGN2 §3 "listas de tarjetas"; spec `visual-design-
    system`, requirement 'Eight DESIGN2 Component Classes')."""
    a = usuario_factory(username="mis-reportes-barra-lista")
    reporte_factory(creador=a)
    client.force_login(a)

    response = client.get(reverse("reportes_mis"))
    contenido = response.content.decode()

    assert response.status_code == 200
    assert 'class="barra-pantalla' in contenido
    assert 'class="lista' in contenido


# ---------------------------------------------------------------------------
# sincronizacion (S-15 aggregated pending/failed sync screen; change
# `vista-sincronizacion-pendientes`, Phase 3, design D1/D3; spec
# `sincronizacion-pendientes`)
# ---------------------------------------------------------------------------


@pytest.mark.django_db
def test_reportes_sincronizacion_resuelve_y_responde_ok(client, usuario_factory):
    """Change `vista-sincronizacion-pendientes`, task 3.1/3.2 (design D1) —
    the route MUST exist and render for an authenticated user with zero ORM
    queries against `Reporte` (the list is built entirely client-side from
    Dexie)."""
    usuario = usuario_factory(username="sincronizacion-resuelve")
    client.force_login(usuario)

    response = client.get(reverse("reportes_sincronizacion"))

    assert response.status_code == 200


def test_reportes_sincronizacion_anonimo_redirige_a_login(client):
    """Change `vista-sincronizacion-pendientes`, task 3.1 (design D1) —
    same `login_required` contract as every other reportes route."""
    response = client.get(reverse("reportes_sincronizacion"))

    assert response.status_code == 302
    assert reverse("login") in response.url


@pytest.mark.django_db
def test_reportes_sincronizacion_expone_hooks_de_shell(client, usuario_factory):
    """Change `vista-sincronizacion-pendientes`, task 3.3/3.4 (design D3) —
    the shell MUST render a CSRF token input, the `{% url 'reportes_paso' 0
    '__SECCION__' %}` retry-URL placeholder, and the list/empty-state
    hooks the client JS (Phase 4) will bind against."""
    usuario = usuario_factory(username="sincronizacion-hooks")
    client.force_login(usuario)

    response = client.get(reverse("reportes_sincronizacion"))
    contenido = response.content.decode()

    assert response.status_code == 200
    assert "csrfmiddlewaretoken" in contenido
    assert reverse("reportes_paso", args=[0, "__SECCION__"]) in contenido
    assert "data-sincronizacion-lista" in contenido
    assert "data-sincronizacion-vacio" in contenido


@pytest.mark.django_db
def test_mis_reportes_expone_badge_de_pendientes(client, usuario_factory):
    """Change `vista-sincronizacion-pendientes`, task 5.1/5.2 (design D5) —
    `mis_reportes.html` MUST render a hidden badge link to
    `reportes_sincronizacion` and load `offline-db.js`/`pendientes-badge.js`
    so the client can reveal it with a live pending/failed count (spec
    'Entry Point From Mis Reportes')."""
    usuario = usuario_factory(username="mis-reportes-badge")
    client.force_login(usuario)

    response = client.get(reverse("reportes_mis"))
    contenido = response.content.decode()

    assert response.status_code == 200
    assert reverse("reportes_sincronizacion") in contenido
    assert "data-badge-pendientes" in contenido
    assert "reportes/offline-db.js" in contenido
    assert "reportes/pendientes-badge.js" in contenido


# ---------------------------------------------------------------------------
# seleccion_de_tipo (S-03; change `mis-reportes-agrupado-por-estado` Phase 5;
# spec `seleccion-tipo-reporte`; design D6). Entry point reached from
# "Mis reportes" (S-02) "+ Nuevo reporte", submitting to the existing,
# untouched `reportes_nuevo` route.
# ---------------------------------------------------------------------------


@pytest.mark.django_db
def test_seleccion_de_tipo_lista_activos(
    client, usuario_factory, tipo_con_definicion_activa_factory
):
    """Spec 'Active Tipo De Reporte Listing' — 'Active types are listed':
    two active `TipoDeReporte` rows with distinct códigos are both listed
    with their código and section count."""
    tipo_a, _definicion_a = tipo_con_definicion_activa_factory(
        codigo="seleccion-tipo-activo-a"
    )
    tipo_b, _definicion_b = tipo_con_definicion_activa_factory(
        codigo="seleccion-tipo-activo-b"
    )
    usuario = usuario_factory(username="seleccion-tipo-lista-activos")
    client.force_login(usuario)

    response = client.get(reverse("reportes_seleccion_tipo"))
    contenido = response.content.decode()

    assert response.status_code == 200
    assert tipo_a.codigo in contenido
    assert tipo_b.codigo in contenido
    # Both fixture tipos share `definicion_valida`'s 2-section estructura.
    assert contenido.count("2 secciones") == 2


@pytest.mark.django_db
def test_seleccion_de_tipo_muestra_inactivos_deshabilitados(
    client, usuario_factory
):
    """Spec 'Inactive Types Shown Disabled' — 'Inactive type cannot be
    selected': a `TipoDeReporte` with no `definicion_activa` (design D6:
    `activo` is a property, never a queryable column) appears disabled with
    a "próximamente" label and its selection control does not submit."""
    tipo_inactivo = TipoDeReporte.objects.create(
        nombre="Tipo sin activar",
        codigo="seleccion-tipo-inactivo",
        plantilla=SimpleUploadedFile(
            "plantilla.xlsx", b"contenido-irrelevante-para-este-nivel"
        ),
    )
    usuario = usuario_factory(username="seleccion-tipo-inactivo-usuario")
    client.force_login(usuario)

    response = client.get(reverse("reportes_seleccion_tipo"))
    contenido = response.content.decode()

    assert response.status_code == 200
    assert tipo_inactivo.codigo in contenido
    assert "próximamente" in contenido
    assert f'action="{reverse("reportes_nuevo", args=[tipo_inactivo.codigo])}"' not in contenido


@pytest.mark.django_db
def test_seleccion_de_tipo_anonimo_redirige(client):
    """Spec 'Active Tipo De Reporte Listing' — 'Anonymous user is
    redirected': no authenticated session redirects to the login flow."""
    response = client.get(reverse("reportes_seleccion_tipo"))

    assert response.status_code == 302
    assert reverse("login") in response.url


@pytest.mark.django_db
def test_seleccion_de_tipo_selecciona_activo_crea_reporte(
    client, usuario_factory, tipo_con_definicion_activa_factory
):
    """Spec 'Submits To Existing Nuevo Reporte Route' — 'Selecting an
    active type creates a report': the S-03 screen's form for an active
    tipo posts to the existing `reportes_nuevo` route, unchanged, creating
    the `Reporte` via that route's own logic — this screen duplicates no
    creation logic."""
    tipo, definicion = tipo_con_definicion_activa_factory(
        codigo="seleccion-tipo-selecciona-activo"
    )
    usuario = usuario_factory(username="seleccion-tipo-selecciona-activo-usuario")
    client.force_login(usuario)

    respuesta_lista = client.get(reverse("reportes_seleccion_tipo"))
    contenido_lista = respuesta_lista.content.decode()

    assert respuesta_lista.status_code == 200
    assert f'action="{reverse("reportes_nuevo", args=[tipo.codigo])}"' in contenido_lista
    assert "csrfmiddlewaretoken" in contenido_lista

    response = client.post(reverse("reportes_nuevo", args=[tipo.codigo]))

    assert response.status_code == 302
    assert Reporte.objects.filter(tipo=tipo, creador=usuario).count() == 1


@pytest.mark.django_db
def test_sidebar_enlaza_a_sincronizacion(client, usuario_factory):
    """Hueco de navegación detectado en la verificación manual del change
    `vista-sincronizacion-pendientes`: el service worker cachea S-15, pero su
    único enlace vivía en `mis_reportes.html`, que NO está cacheada. Sin
    conexión, alguien con pasos pendientes no tenía forma de llegar a la
    pantalla que los resuelve salvo escribir la URL de memoria.

    El sidebar se renderiza desde `base.html` en TODA pantalla autenticada,
    incluidas las que el service worker cachea, así que el enlace viaja con
    ellas y queda disponible con o sin conexión."""
    usuario = usuario_factory(username="ve-sincronizacion")
    client.force_login(usuario)

    respuesta = client.get(reverse("reportes_mis"))
    contenido = respuesta.content.decode()

    inicio_sidebar = contenido.index('class="escritorio__sidebar"')
    fin_sidebar = contenido.index("</nav>", inicio_sidebar)
    sidebar = contenido[inicio_sidebar:fin_sidebar]

    assert reverse("reportes_sincronizacion") in sidebar

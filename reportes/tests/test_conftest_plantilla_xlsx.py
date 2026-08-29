"""Test for the `plantilla_xlsx` fixture added to `reportes/tests/conftest.py`
(backlog #7, design's Testing Strategy — "app-local duplication is this
repo's stated convention"). Mirrors `tipos_reporte/tests/conftest.py`'s
fixture of the same name and shape.

Strict TDD: written RED (referencing a fixture that does not exist yet)
before the fixture lands in `conftest.py`.
"""

from openpyxl import load_workbook


def test_plantilla_xlsx_construye_libro_real_con_hoja_y_rangos(plantilla_xlsx):
    destino = plantilla_xlsx(
        nombre_hoja="REPORTE", rangos=("M10:P10", "M12:P12", "M25:P25")
    )

    libro = load_workbook(destino)

    assert libro.sheetnames == ["REPORTE"]
    rangos_fusionados = {str(rango) for rango in libro["REPORTE"].merged_cells.ranges}
    assert rangos_fusionados == {"M10:P10", "M12:P12", "M25:P25"}

"""Fixtures for the `reportes` app test suite (backlog #5).

App-local duplication of `usuario_factory`/`definicion_valida` is the
established project convention — `tipos_reporte` already duplicates
`usuario_factory` from `usuarios` deliberately (design's Testing Strategy).
"""

import copy

import openpyxl
import pytest
from django.core.files.uploadedfile import SimpleUploadedFile
from django.utils import timezone


@pytest.fixture
def usuario_factory(db):
    """Create a Usuario via `create_user`, so passwords hash (design's
    Testing Strategy — distinct from `tipos_reporte`'s plaintext-password
    fixture because `cliente_autenticado` needs a real login)."""
    from usuarios.models import Usuario

    def _create(**kwargs):
        defaults = {
            "username": "usuario_test",
            "password": "contrasena-de-prueba-123",
        }
        defaults.update(kwargs)
        return Usuario.objects.create_user(**defaults)

    return _create


@pytest.fixture
def definicion_valida():
    """A minimal but complete `estructura` dict: one section with one campo
    and one item, both fully specified (mirrored from
    `tipos_reporte/tests/conftest.py`, deep-copied per call)."""

    def _crear():
        return copy.deepcopy(
            {
                "tipo": "instalacion-resinas",
                "plantilla": "JME.PC-0001.F1.xlsx",
                "hoja": "REPORTE",
                "secciones": [
                    {
                        "id": "datos-generales",
                        "titulo": "Datos generales",
                        "campos": [
                            {
                                "id": "turno",
                                "etiqueta": "Turno",
                                "tipo": "seleccion",
                                "opciones": ["Día", "Noche"],
                                "obligatorio": True,
                                "celda": "M12",
                            }
                        ],
                    },
                    {
                        "id": "proceso-instalacion",
                        "titulo": "Proceso de instalación",
                        "roles": ["construccion-jme", "qa-subterra"],
                        "items": [
                            {
                                "id": "p-01",
                                "texto": "Se verifica ángulo de perforación.",
                                "tipo": "rango-hora-inicio-fin",
                                "celda_inicio": "M25",
                                "celda_fin": "P25",
                            }
                        ],
                    },
                ],
            }
        )

    return _crear


@pytest.fixture
def estructura_con_validaciones():
    """An `estructura` dict exercising every rule `reportes.validacion`
    checks (backlog `validacion-datos-formulario`, spec scenarios 1-6): one
    `obligatorio` `texto` campo, one `obligatorio` `seleccion` campo whose
    `opciones` include `"No cumple"`, and one `obligatorio`
    `rango-hora-inicio-fin` item — deep-copied per call, mirroring
    `definicion_valida`."""

    def _crear():
        return copy.deepcopy(
            {
                "tipo": "instalacion-resinas",
                "plantilla": "JME.PC-0001.F1.xlsx",
                "hoja": "REPORTE",
                "secciones": [
                    {
                        "id": "datos-generales",
                        "titulo": "Datos generales",
                        "campos": [
                            {
                                "id": "observaciones-generales",
                                "etiqueta": "Observaciones generales",
                                "tipo": "texto",
                                "obligatorio": True,
                                "celda": "M10",
                            },
                            {
                                "id": "estado-general",
                                "etiqueta": "Estado general",
                                "tipo": "seleccion",
                                "opciones": ["Cumple", "No cumple"],
                                "obligatorio": True,
                                "celda": "M12",
                            },
                        ],
                    },
                    {
                        "id": "proceso-instalacion",
                        "titulo": "Proceso de instalación",
                        "items": [
                            {
                                "id": "p-01",
                                "texto": "Se verifica ángulo de perforación.",
                                "tipo": "rango-hora-inicio-fin",
                                "obligatorio": True,
                                "celda_inicio": "M25",
                                "celda_fin": "P25",
                            }
                        ],
                    },
                ],
            }
        )

    return _crear


@pytest.fixture
def tipo_con_definicion_activa_factory(db, definicion_valida):
    """Create a `TipoDeReporte` with an already-activated `DefinicionDeTipo`
    (design D11): the row is built directly in activated shape
    (`estado=ACTIVA`, `version=1`, `activada_en=now`) to satisfy the
    `definicion_estado_implica_version` CheckConstraint, bypassing
    `servicios.activar_definicion` — activation itself is #3's tested
    responsibility, not the wizard's. Returns `(tipo, definicion)`."""
    from tipos_reporte.models import DefinicionDeTipo, Estado, TipoDeReporte

    def _crear(**kwargs):
        estructura = kwargs.pop("estructura", None) or definicion_valida()
        tipo = TipoDeReporte.objects.create(
            nombre=kwargs.pop("nombre", "Instalación de resinas"),
            codigo=kwargs.pop("codigo", "instalacion-resinas"),
            plantilla=kwargs.pop(
                "plantilla",
                SimpleUploadedFile(
                    "plantilla.xlsx", b"contenido-irrelevante-para-este-nivel"
                ),
            ),
        )
        definicion = DefinicionDeTipo.objects.create(
            tipo=tipo,
            archivo_yaml=SimpleUploadedFile(
                "definicion.yaml", b"secciones: []"
            ),
            yaml_fuente="secciones: []",
            estructura=estructura,
            estado=Estado.ACTIVA,
            version=1,
            activada_en=timezone.now(),
        )
        tipo.definicion_activa = definicion
        tipo.save(update_fields=["definicion_activa"])
        return tipo, definicion

    return _crear


@pytest.fixture
def reporte_factory(db, usuario_factory, tipo_con_definicion_activa_factory):
    """Create a `Reporte` with sensible defaults, overridable by kwargs."""
    from reportes.models import Reporte

    def _crear(**kwargs):
        if "tipo" not in kwargs or "definicion" not in kwargs:
            tipo, definicion = tipo_con_definicion_activa_factory()
            kwargs.setdefault("tipo", tipo)
            kwargs.setdefault("definicion", definicion)
        # `setdefault(key, usuario_factory())` would call `usuario_factory()`
        # eagerly on every invocation regardless of whether `creador` is
        # already in `kwargs` (Python evaluates arguments before the call),
        # creating an unwanted extra Usuario row that can collide with an
        # already-created "usuario_test" from another fixture in the same
        # test (e.g. `cliente_autenticado`). Only call it when actually
        # needed.
        if "creador" not in kwargs:
            kwargs["creador"] = usuario_factory()
        return Reporte.objects.create(**kwargs)

    return _crear


@pytest.fixture
def plantilla_xlsx(tmp_path):
    """A real `.xlsx` workbook built with openpyxl, not committed to the
    repository (design's Testing Strategy — mirrors
    `tipos_reporte/tests/conftest.py`'s fixture of the same name; app-local
    duplication is this repo's stated convention). By default declares one
    sheet named "REPORTE" with one merged range `M12:P12`. Generation tests
    (backlog #7) need `rangos=("M10:P10", "M12:P12", "M25:P25")`, since
    `tipo_con_definicion_activa_factory`'s default `plantilla` upload
    (`b"contenido-irrelevante-para-este-nivel"`) cannot be parsed by
    `load_workbook`. Returns a factory so each test can vary the sheet name
    and merged ranges."""

    def _crear(nombre_hoja="REPORTE", rangos=("M12:P12",)):
        wb = openpyxl.Workbook()
        wb.active.title = nombre_hoja
        for rango in rangos:
            wb.active.merge_cells(rango)
        destino = tmp_path / "plantilla.xlsx"
        wb.save(destino)
        return destino

    return _crear


@pytest.fixture
def reporte_listo_para_cerrar(
    client,
    usuario_factory,
    estructura_con_validaciones,
    tipo_con_definicion_activa_factory,
    plantilla_xlsx,
):
    """A `(client, reporte)` pair ready for `cerrar_reporte` and `generar`
    (backlog #7, tasks 3.2/5, design's Testing Strategy): built on
    `estructura_con_validaciones` with a real `.xlsx` template and all four
    obligatorio `ValorDeReporte` rows persisted, so `puede_generar` is
    True. The creador is already logged into `client`.

    `estructura_con_validaciones`'s `p-01` item declares TWO independent
    anchor cells on the same row (`celda_inicio="M25"`,
    `celda_fin="P25"`) — a single `M25:P25` merge would swallow `P25` into
    a read-only `MergedCell` (openpyxl only the merge's top-left cell is
    writable), so `M25` and `P25` are deliberately left as two ordinary,
    unmerged, independently-writable cells here."""
    from reportes.models import Reporte, ValorDeReporte

    destino = plantilla_xlsx(rangos=("M10:P10", "M12:P12"))
    with open(destino, "rb") as archivo:
        contenido = archivo.read()

    creador = usuario_factory(username="creador_listo_para_cerrar")
    tipo, definicion = tipo_con_definicion_activa_factory(
        estructura=estructura_con_validaciones(),
        plantilla=SimpleUploadedFile("plantilla.xlsx", contenido),
    )
    reporte = Reporte.objects.create(
        tipo=tipo, definicion=definicion, creador=creador
    )

    for identificador, valor in (
        ("observaciones-generales", "Todo en orden."),
        ("estado-general", "Cumple"),
        ("p-01_inicio", "08:00"),
        ("p-01_fin", "09:00"),
    ):
        ValorDeReporte.objects.create(
            reporte=reporte,
            identificador_de_campo=identificador,
            valor=valor,
            autor=creador,
        )

    client.force_login(creador)
    return client, reporte


@pytest.fixture
def participacion_factory(db, usuario_factory):
    """Create a `Usuario` AND its `ParticipacionEnReporte` row on `reporte`
    (backlog #8, design's Fixture strategy), returning the invited user.
    `username` is explicit per call, following the existing convention that
    avoids `"usuario_test"` collisions with `cliente_autenticado`."""
    from reportes.models import ParticipacionEnReporte

    def _crear(reporte, username="invitado"):
        usuario = usuario_factory(username=username)
        ParticipacionEnReporte.objects.create(reporte=reporte, usuario=usuario)
        return usuario

    return _crear


@pytest.fixture
def cliente_autenticado(client, usuario_factory):
    """A Django test client already logged in as a fresh Usuario
    (`client.force_login` — authentication itself is #2's tested concern,
    per design's Testing Strategy)."""
    usuario = usuario_factory()
    client.force_login(usuario)
    return client

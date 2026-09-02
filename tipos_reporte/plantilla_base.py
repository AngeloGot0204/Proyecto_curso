"""Construye una plantilla .xlsx limpia para el formato JME.PC-0001.F1.

Reproduce la estructura visible del formato oficial, pero autorada desde cero
con openpyxl: sin customXml, sin printerSettings binarios, sin imagenes WMF y
sin partes heredadas que openpyxl no sepa reserializar. El archivo que openpyxl
escribe es el mismo que openpyxl vuelve a leer, que es justo lo que rompia con
la plantilla original.

Toda celda destinada a recibir un valor es ancla de su rango combinado (regla
R6 del validador), de modo que el generador nunca escriba en una celda no-ancla.
"""

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

HOJA = "JME.PC-0001.F1"

ANCHOS = {
    "A": 2.0, "B": 9.1, "C": 24.0, "D": 3.3, "E": 6.7, "F": 9.0, "G": 10.0,
    "H": 14.8, "I": 20.8, "J": 9.0, "K": 18.1, "L": 14.1, "M": 7.4, "N": 2.7,
    "O": 10.0, "P": 7.4, "Q": 7.8, "R": 6.7, "S": 13.1, "T": 6.7, "U": 3.0,
    "V": 11.0, "W": 5.4,
}

TINTA = "FF14130F"
LINEA = "FFB0ADA8"
GRIS = "FFEDEBE8"

fina = Side(style="thin", color=LINEA)
media = Side(style="medium", color=TINTA)
BORDE = Border(left=fina, right=fina, top=fina, bottom=fina)
BORDE_FUERTE = Border(left=media, right=media, top=media, bottom=media)

CENTRO = Alignment(horizontal="center", vertical="center", wrap_text=True)
IZQ = Alignment(horizontal="left", vertical="center", wrap_text=True)

NEGRITA = Font(name="Arial", size=9, bold=True, color=TINTA)
NORMAL = Font(name="Arial", size=9, color=TINTA)
TITULO = Font(name="Arial", size=11, bold=True, color=TINTA)


def celda(hoja, coord, valor=None, fuente=NORMAL, alineacion=IZQ,
          borde=BORDE, relleno=None):
    c = hoja[coord]
    if valor is not None:
        c.value = valor
    c.font = fuente
    c.alignment = alineacion
    c.border = borde
    if relleno:
        c.fill = PatternFill("solid", fgColor=relleno)
    return c


def combinar(hoja, rango, **kwargs):
    """Combina `rango` y estiliza TODAS sus celdas.

    Estilar solo el ancla deja el resto del rango sin borde: Excel dibuja el
    recuadro a partir de cada celda, no del rango."""
    hoja.merge_cells(rango)
    ancla = rango.split(":")[0]
    for fila in hoja[rango]:
        for c in fila:
            c.border = kwargs.get("borde", BORDE)
            if kwargs.get("relleno"):
                c.fill = PatternFill("solid", fgColor=kwargs["relleno"])
    return celda(hoja, ancla, **kwargs)


def encabezado_seccion(hoja, fila, titulo, columnas):
    """Barra de sección: título a la izquierda y las columnas de rol."""
    combinar(hoja, f"B{fila}:L{fila}", valor=titulo, fuente=NEGRITA,
             alineacion=IZQ, relleno=GRIS, borde=BORDE_FUERTE)
    for rango, etiqueta in columnas:
        combinar(hoja, f"{rango}{fila}:{rango}{fila}" if ":" not in rango
                 else f"{rango.split(':')[0]}{fila}:{rango.split(':')[1]}{fila}",
                 valor=etiqueta, fuente=NEGRITA, alineacion=CENTRO,
                 relleno=GRIS, borde=BORDE_FUERTE)


def construir():
    wb = Workbook()
    h = wb.active
    h.title = HOJA

    for col, ancho in ANCHOS.items():
        h.column_dimensions[col].width = ancho

    h.page_setup.orientation = "portrait"
    h.page_setup.fitToPage = True
    h.page_setup.fitToWidth = 1
    h.page_setup.fitToHeight = 0
    h.sheet_properties.pageSetUpPr.fitToPage = True
    h.print_area = "B2:V65"

    # ---- Encabezado institucional (2-5) --------------------------------
    combinar(h, "B2:D5", valor="LOGO", fuente=NEGRITA, alineacion=CENTRO,
             borde=BORDE_FUERTE)
    combinar(h, "E2:Q2", valor="REGISTRO", fuente=NEGRITA, alineacion=CENTRO,
             borde=BORDE_FUERTE)
    combinar(h, "E3:Q3", valor="AREA DE CALIDAD", fuente=NEGRITA,
             alineacion=CENTRO, borde=BORDE_FUERTE)
    combinar(h, "E4:Q5",
             valor="REPORTE DE VERIFICACION DE INSTALACION DE PERNOS DE "
                   "ANCLAJES CON RESINA",
             fuente=TITULO, alineacion=CENTRO, borde=BORDE_FUERTE)
    combinar(h, "R2:V2", valor="JME.SGC.18138.PC-0001-F1", fuente=NORMAL,
             alineacion=CENTRO, borde=BORDE_FUERTE)
    combinar(h, "R3:V3", valor="Revisión: 0", alineacion=CENTRO,
             borde=BORDE_FUERTE)
    combinar(h, "R4:V4", valor="Fecha: 01 / 02 / 2020", alineacion=CENTRO,
             borde=BORDE_FUERTE)
    combinar(h, "R5:V5", valor="Página:", alineacion=CENTRO,
             borde=BORDE_FUERTE)

    # ---- Identificación (6-13) -----------------------------------------
    # (etiqueta, rango_etiqueta, rango_valor)
    identificacion = [
        ("NOMBRE DEL PROYECTO:", "B6:D6", "E6:Q6"),
        ("N° REGISTRO:", "R6:S6", "T6:V6"),
        ("COMPAÑIA:", "B7:D7", "E7:I7"),
        ("N° CONTRATO:", "J7:L7", "M7:Q7"),
        ("FECHA:", "R7:S7", "T7:V7"),
        ("PLANO REF:", "B9:D9", "E9:I9"),
        ("AREA:", "J9:L9", "M9:O9"),
        ("FRENTE:", "P9:Q9", "R9:V9"),
        ("SISTEMA:", "B10:D10", "E10:I10"),
        ("SUB-SISTEMA:", "J10:L10", "M10:V10"),
        ("PPI:", "B11:D11", "E11:V11"),
        ("COD. DISEÑO:", "B12:D12", "E12:I12"),
        ("TURNO:", "J12:L12", "M12:O12"),
        ("TIPO ROCA:", "P12:Q12", "R12:V12"),
        ("DESCRIPCION:", "B13:D13", "E13:V13"),
    ]
    for etiqueta, rango_etq, rango_val in identificacion:
        combinar(h, rango_etq, valor=etiqueta, fuente=NEGRITA, alineacion=IZQ)
        combinar(h, rango_val)

    # ---- Parámetros preliminares (15-22) -------------------------------
    combinar(h, "B15:L15", valor="PARAMETROS PRELIMINARES", fuente=NEGRITA,
             alineacion=IZQ, relleno=GRIS, borde=BORDE_FUERTE)
    combinar(h, "M15:O15", valor="CONSORCIO JME", fuente=NEGRITA,
             alineacion=CENTRO, relleno=GRIS, borde=BORDE_FUERTE)
    combinar(h, "P15:R15", valor="QA SUBTERRA", fuente=NEGRITA,
             alineacion=CENTRO, relleno=GRIS, borde=BORDE_FUERTE)
    combinar(h, "S15:V15", valor="OBSERVACION", fuente=NEGRITA,
             alineacion=CENTRO, relleno=GRIS, borde=BORDE_FUERTE)

    preliminares = [
        "1.- Se verifica operatividad de equipo.",
        "2.- Se tiene resinas en buen estado.",
        "3.- Se verifica limpieza, diámetro y longitud del perno.",
        "4.- Se verifica el diámetro de la broca de equipo Bolter.",
        "5.- Se verifica longitud de los taladros usados en la perforación.",
        "6.- Se verifica certificado de calibración del equipo pull test.",
        "7.- Se verifica certificado de calidad de los materiales.",
    ]
    for i, texto in enumerate(preliminares):
        f = 16 + i
        h.row_dimensions[f].height = 22
        combinar(h, f"B{f}:L{f}", valor=texto, alineacion=IZQ)
        combinar(h, f"M{f}:O{f}", alineacion=CENTRO)
        combinar(h, f"P{f}:R{f}", alineacion=CENTRO)
        combinar(h, f"S{f}:V{f}")

    # ---- Proceso de instalación (24-36) --------------------------------
    combinar(h, "B24:L24",
             valor="PROCESO DE INSTALACION DE RESINAS PARA PERNOS DE ANCLAJES",
             fuente=NEGRITA, alineacion=IZQ, relleno=GRIS, borde=BORDE_FUERTE)
    combinar(h, "M24:O24", valor="Hora inicio", fuente=NEGRITA,
             alineacion=CENTRO, relleno=GRIS, borde=BORDE_FUERTE)
    combinar(h, "P24:R24", valor="Hora término", fuente=NEGRITA,
             alineacion=CENTRO, relleno=GRIS, borde=BORDE_FUERTE)
    combinar(h, "S24:V24", valor="OBSERVACION", fuente=NEGRITA,
             alineacion=CENTRO, relleno=GRIS, borde=BORDE_FUERTE)

    proceso = [
        "Se verifica ángulo de perforación.",
        "El equipo inició y culminó la perforación.",
        "Se verifica la limpieza y longitud de perforación ejecutada.",
        "Se ingresan las resinas acorde al diámetro y perforación del taladro.",
        "Diámetro de resina.",
        "Longitud de cartucho de resina.",
        "Número de resinas insertadas.",
        "Se ocupó totalmente la perforación con las resinas.",
        "Se coloca el perno en el equipo.",
        "Se introduce el perno por rotación.",
        "Inicio y culminación del mezclado de resinas.",
    ]
    for i, texto in enumerate(proceso):
        f = 25 + i
        h.row_dimensions[f].height = 22
        celda(h, f"B{f}", valor=str(i + 1), fuente=NEGRITA, alineacion=CENTRO)
        combinar(h, f"C{f}:L{f}", valor=texto, alineacion=IZQ)
        combinar(h, f"M{f}:O{f}", alineacion=CENTRO)
        combinar(h, f"P{f}:R{f}", alineacion=CENTRO)
        combinar(h, f"S{f}:V{f}")

    # ---- Ensayo de Pull Test (37-42) -----------------------------------
    combinar(h, "B37:L37", valor="ENSAYO DE PULL TEST", fuente=NEGRITA,
             alineacion=IZQ, relleno=GRIS, borde=BORDE_FUERTE)
    combinar(h, "M37:O37", valor="Hora inicio", fuente=NEGRITA,
             alineacion=CENTRO, relleno=GRIS, borde=BORDE_FUERTE)
    combinar(h, "P37:R37", valor="Hora término", fuente=NEGRITA,
             alineacion=CENTRO, relleno=GRIS, borde=BORDE_FUERTE)
    combinar(h, "S37:V37", valor="OBSERVACION", fuente=NEGRITA,
             alineacion=CENTRO, relleno=GRIS, borde=BORDE_FUERTE)

    pull_test = [
        ("12", "Se inicia el proceso de instalación del equipo Pull Test."),
        ("13", "Se inicia el proceso de aplicación de tracción del perno."),
        ("", "Fuerza de tracción aplicada."),
        ("14", "Tiempo transcurrido de tensión."),
        ("15", "Desplazamiento del perno."),
    ]
    for i, (numero, texto) in enumerate(pull_test):
        f = 38 + i
        h.row_dimensions[f].height = 22
        celda(h, f"B{f}", valor=numero, fuente=NEGRITA, alineacion=CENTRO)
        combinar(h, f"C{f}:L{f}", valor=texto, alineacion=IZQ)
        combinar(h, f"M{f}:O{f}", alineacion=CENTRO)
        combinar(h, f"P{f}:R{f}", alineacion=CENTRO)
        combinar(h, f"S{f}:V{f}")

    # ---- Resultados (44-48) --------------------------------------------
    combinar(h, "B44:L44", valor="RESULTADOS DE LA PRUEBA", fuente=NEGRITA,
             alineacion=IZQ, relleno=GRIS, borde=BORDE_FUERTE)
    combinar(h, "M44:O44", valor="QC JME", fuente=NEGRITA, alineacion=CENTRO,
             relleno=GRIS, borde=BORDE_FUERTE)
    combinar(h, "P44:R44", valor="QA Antamina", fuente=NEGRITA,
             alineacion=CENTRO, relleno=GRIS, borde=BORDE_FUERTE)
    combinar(h, "S44:V44", valor="OBSERVACION", fuente=NEGRITA,
             alineacion=CENTRO, relleno=GRIS, borde=BORDE_FUERTE)

    resultados = [
        "El mezclado de la resina fue en el tiempo acorde a la hoja técnica.",
        "El resultado del ensayo a tracción fue aceptable.",
        "El tiempo de fragua es aceptable.",
    ]
    for i, texto in enumerate(resultados):
        f = 45 + i
        h.row_dimensions[f].height = 22
        celda(h, f"B{f}", valor=str(i + 1), fuente=NEGRITA, alineacion=CENTRO)
        combinar(h, f"C{f}:L{f}", valor=texto, alineacion=IZQ)
        combinar(h, f"M{f}:O{f}", alineacion=CENTRO)
        combinar(h, f"P{f}:R{f}", alineacion=CENTRO)
        combinar(h, f"S{f}:V{f}")

    # ---- Croquis y observaciones (50-58) -------------------------------
    combinar(h, "B50:L50", valor="CROQUIS DE ZONA DE TRABAJO:", fuente=NEGRITA,
             alineacion=IZQ, relleno=GRIS, borde=BORDE_FUERTE)
    combinar(h, "M50:V50", valor="OBSERVACIONES:", fuente=NEGRITA,
             alineacion=IZQ, relleno=GRIS, borde=BORDE_FUERTE)
    for f in range(51, 59):
        h.row_dimensions[f].height = 18
    combinar(h, "B51:L58", borde=BORDE_FUERTE,
             alineacion=Alignment(horizontal="center", vertical="center"))
    combinar(h, "M51:V58", borde=BORDE_FUERTE,
             alineacion=Alignment(horizontal="left", vertical="top",
                                  wrap_text=True))

    # ---- Firmas (60-65) -------------------------------------------------
    roles = [
        ("Construcción JME.", "B", "F"),
        ("QC JME.", "G", "I"),
        ("QA SUBTERRA.", "J", "N"),
        ("Construcción Antamina.", "O", "V"),
    ]
    for titulo, desde, hasta in roles:
        combinar(h, f"{desde}60:{hasta}60", valor=titulo, fuente=NEGRITA,
                 alineacion=CENTRO, relleno=GRIS, borde=BORDE_FUERTE)
        h.row_dimensions[61] = h.row_dimensions[61]
        combinar(h, f"{desde}61:{hasta}63", valor="Firma:", alineacion=IZQ,
                 borde=BORDE_FUERTE)
        combinar(h, f"{desde}64:{hasta}64", valor="Nombre:", alineacion=IZQ)
        combinar(h, f"{desde}65:{hasta}65", valor="Fecha:", alineacion=IZQ)

    h.row_dimensions[61].height = 30
    h.row_dimensions[62].height = 18
    h.row_dimensions[63].height = 18

    return wb

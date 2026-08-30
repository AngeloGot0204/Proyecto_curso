"""Report generation service for `tipos_reporte` (backlog #4).

`generar_reporte(definicion, valores)` fills an already-**activated**
`DefinicionDeTipo`'s template with captured values and returns the
generated `.xlsx` bytes for the single declared sheet. It mutates a
*loaded copy* of the original template — never the stored file, never a
rebuilt workbook (ADR-0002) — and reuses `tipos_reporte/validacion.py`'s
`_iterar_nodos`/`_claves_de_celda_requeridas` so anchor logic exists in
exactly one place (design D1).

This module implements every phase of the change's `tasks.md`: Phase 1
(exceptions), Phase 2 (template loading + completeness validation), Phase 3
(cell writing + sheet-only export) and Phase 4 (logo swap).
"""

import logging
from io import BytesIO

from openpyxl import load_workbook
from openpyxl.drawing.image import Image as ImagenOpenpyxl

from tipos_reporte.validacion import _claves_de_celda_requeridas, _iterar_nodos

logger = logging.getLogger(__name__)

# Design D5's Open Questions default (confirmed for this PR): a declared
# attachment anchor slot with no explicit `ancho_px`/`alto_px` fits the
# image within a 320x240 box, aspect ratio preserved.
_ANCHO_PX_POR_DEFECTO = 320
_ALTO_PX_POR_DEFECTO = 240


class ProblemaDeGeneracion(Exception):
    """Base for every foreseeable generation failure (design's
    Interfaces/Contracts), so backlog #7's future endpoint catches one type
    and never surfaces a raw 500 for a foreseeable problem."""

    regla = "problema-de-generacion"


class PlantillaIlegible(ProblemaDeGeneracion):
    """The template file could not be opened or parsed as a workbook.
    Same stable `regla` id `validacion.validar_contra_plantilla` uses for
    the equivalent problem, per design's Interfaces/Contracts."""

    regla = "plantilla-ilegible"


class ValoresIncompletos(ProblemaDeGeneracion):
    """One or more required ids are missing from `valores` (design D2/D3).
    `faltantes` is the stable, test-assertable identifier — the message is
    free to be reworded, mirroring `ProblemaDeDefinicion`'s convention
    (`validacion.py:40-50`). Every missing id is accumulated into one raise
    (never fail-fast)."""

    regla = "valores-incompletos"

    def __init__(self, faltantes):
        self.faltantes = tuple(sorted(faltantes))
        super().__init__(
            "Faltan valores obligatorios para generar el reporte: "
            + ", ".join(self.faltantes)
        )


# Design D1: the suffix a `valores` key gets, keyed by which cell key the
# node declares — "" for a scalar `celda`, "_inicio"/"_fin" for a range's
# two independent keys.
_SUFIJO_POR_CLAVE = {"celda": "", "celda_inicio": "_inicio", "celda_fin": "_fin"}


def claves_de_valor(nodo):
    """`valores`-dict keys one campo/item declares (design D5). Public so
    both this module and `reportes` (the wizard app) derive the exact same
    `ValorDeReporte.identificador_de_campo` names from one owner — never a
    copy of `_SUFIJO_POR_CLAVE` that could drift. Extracted from
    `_destinos`, which still owns pairing each key with its cell
    coordinate."""
    return [
        f"{nodo['id']}{_SUFIJO_POR_CLAVE[clave]}"
        for clave in _claves_de_celda_requeridas(nodo.get("tipo"))
    ]


def _destinos(nodo):
    """`(clave_en_valores, coordenada)` pairs for one campo/item (design
    D1). Single derivation shared by the completeness pass and the write
    pass, so a validated-as-present key can never be written to the wrong
    cell — `_claves_de_celda_requeridas` already encodes range-vs-scalar.
    Keys come from `claves_de_valor` (design D5); this function only adds
    each key's cell coordinate."""
    return [
        (clave, nodo[clave_de_celda])
        for clave, clave_de_celda in zip(
            claves_de_valor(nodo), _claves_de_celda_requeridas(nodo.get("tipo"))
        )
    ]


def _validar_completitud(estructura, valores):
    """Accumulate every required id absent from `valores` and raise ONE
    `ValoresIncompletos` listing all of them (design's Sequence, step 4;
    "accumulate every problem"; D2: membership test, never truthiness, so
    `False`/`0`/`""` count as present; D3: only `nodo.get("obligatorio")`
    truthy nodes are required)."""
    faltantes = [
        clave
        for _ubicacion, nodo, _clave_de_etiqueta in _iterar_nodos(estructura)
        if nodo.get("obligatorio")
        for clave, _coordenada in _destinos(nodo)
        if clave not in valores
    ]
    if faltantes:
        raise ValoresIncompletos(faltantes)


def _intercambiar_logo(hoja, logo):
    """Swap the template's original image for `logo` at the SAME anchor
    (design D4, Sequence step 5). Reusing the loaded `anchor` OBJECT (not
    just its coordinates) preserves both position and extent — openpyxl
    serializes an image's box from `anchor._from`/`anchor.ext`, ignoring
    `Image.width`/`height`, so copying pixel dimensions instead would
    resize the drawing box to the new image's size.

    - No `logo` set → leave `hoja._images` untouched (template's original
      logo stays, per the confirmed proposal decision).
    - `logo` set but the template has no image (`hoja._images` empty) →
      no anchor exists to honor, so nothing is inserted (rejected
      alternative: a hardcoded cell, which would drop the logo over report
      data)."""
    originales = hoja._images
    if logo and originales:
        original = originales[0]
        nueva = ImagenOpenpyxl(BytesIO(logo.read()))
        nueva.anchor = original.anchor
        hoja._images.remove(original)
        hoja.add_image(nueva)


def _encajar(imagen, slot):
    """Aspect-preserving fit of `imagen` inside `slot`'s declared box
    (`ancho_px`/`alto_px`, defaulting to `_ANCHO_PX_POR_DEFECTO` x
    `_ALTO_PX_POR_DEFECTO`, design D5). Contrast with `_intercambiar_logo`:
    there a pre-existing anchor OBJECT defines position and extent and
    `Image.width/height` is ignored; here there is no pre-existing drawing,
    so `Image.width/height` is exactly what openpyxl derives the new
    `OneCellAnchor`'s `ext` (extent) from."""
    ancho_max = slot.get("ancho_px", _ANCHO_PX_POR_DEFECTO)
    alto_max = slot.get("alto_px", _ALTO_PX_POR_DEFECTO)
    escala = min(ancho_max / imagen.width, alto_max / imagen.height)
    return round(imagen.width * escala), round(imagen.height * escala)


def _incrustar_adjuntos(hoja, estructura, adjuntos):
    """Embed up to `_MAXIMO_DE_ANCLAS_DE_ADJUNTOS` stored attachments into
    the template's declared anchor slots (backlog #11, design D5, spec
    "Attachment Embedding via Anchor Slots"). A generalization of, not a
    reuse of, `_intercambiar_logo`'s single-fixed-anchor mechanism: this
    uses openpyxl's STRING anchor form (`hoja.add_image(img, "B40")`),
    which makes openpyxl build a fresh `OneCellAnchor` FROM `Image.width/
    height`, the inverse of `_intercambiar_logo`'s object-reuse mechanism.

    `zip` truncation is what enforces the anchor-slot cap: attachments
    beyond the declared slot count are simply never reached, staying
    stored and listable server-side (spec scenario 2); no declared
    `adjuntos` key or no stored attachments both yield zero iterations
    (spec scenarios 3 and 4).

    A file Pillow cannot decode (e.g. an unconverted HEIC that reached
    storage — design's documented risk) is skipped, never raised: turning
    it into a `ProblemaDeGeneracion` would let one attachment block the
    whole document, the same failure mode the "bloqueo solo del adjunto"
    isolation requirement forbids one layer down (design D5)."""
    for slot, archivo in zip(estructura.get("adjuntos") or [], adjuntos):
        try:
            imagen = ImagenOpenpyxl(BytesIO(archivo.read()))
        except Exception:
            logger.exception(
                "No se pudo decodificar un adjunto para incrustarlo en el "
                "reporte generado; se omite y continúa la generación."
            )
            continue
        imagen.width, imagen.height = _encajar(imagen, slot)
        hoja.add_image(imagen, slot["celda"])


def _escribir_valores(hoja, estructura, valores):
    """Write every present `valores` key into its anchor cell (design's
    Sequence, step 6). Walks the same nodes `_validar_completitud` walks,
    but writes ANY key present in `valores` — not only required ones,
    per D3: present-but-not-required keys are written; absent optional
    keys leave their anchor cell untouched ("lo que no se escribe, no se
    altera"). D2: membership test, so `False`/`0`/`""` are written as-is,
    never skipped as if absent."""
    for _ubicacion, nodo, _clave_de_etiqueta in _iterar_nodos(estructura):
        for clave, coordenada in _destinos(nodo):
            if clave in valores:
                hoja[coordenada] = valores[clave]


def _exportar_solo_hoja_declarada(libro, nombre_hoja):
    """Delete every sheet other than the declared one and reset the active
    index (design D5). Deleting sheet OBJECTS never touches the target
    sheet's own merges/styles/drawings, since those live per-sheet or in
    workbook-level shared tables — no rebuild."""
    for nombre in [nombre for nombre in libro.sheetnames if nombre != nombre_hoja]:
        del libro[nombre]
    libro.active = 0


def generar_reporte(definicion, valores: dict, adjuntos=()):
    """Fill `definicion`'s template with `valores` and return the generated
    `.xlsx` bytes (design's Sequence, steps 1-8).

    `definicion` must already be an ACTIVATED `DefinicionDeTipo` — this
    function trusts `definicion.estructura` and adds no re-validation of
    anchor cells (ADR-0002 Compliance table, "Write only to merged-range
    anchor cells").

    `adjuntos` (backlog #11, design D5) is an optional keyword-only-in-
    spirit iterable of file-like objects (each exposing `.read()`) — the
    caller's already-fetched `Adjunto.archivo` values, injected rather than
    queried, so `tipos_reporte` never imports `reportes` (dependency
    direction). Defaults to `()`, so every pre-existing
    `generar_reporte(definicion, valores)` call remains valid unchanged.
    """
    tipo = definicion.tipo
    estructura = definicion.estructura
    plantilla = tipo.plantilla

    try:
        plantilla.open("rb")
    except (OSError, FileNotFoundError) as error:
        # Mirrors `servicios.activar_definicion`'s own guard: a missing
        # file in storage (deleted, ephemeral filesystem) must become a
        # `PlantillaIlegible`, never an uncaught exception (design's
        # Sequence, step 1).
        raise PlantillaIlegible(
            "No se pudo leer el archivo de plantilla (.xlsx)."
        ) from error

    try:
        try:
            libro = load_workbook(plantilla)
        except Exception as error:
            # Any parser failure (not just IO) becomes the same typed
            # problem (design's Sequence, step 2), mirroring
            # `validacion.validar_contra_plantilla`'s convention.
            raise PlantillaIlegible(
                "No se pudo leer el archivo de plantilla (.xlsx)."
            ) from error
    finally:
        plantilla.close()

    nombre_hoja = estructura["hoja"]
    try:
        hoja = libro[nombre_hoja]
    except KeyError as error:
        # Declared sheet missing from the actual workbook (design's
        # Sequence, step 3) — same typed problem, never a raw KeyError.
        raise PlantillaIlegible(
            f"La hoja '{nombre_hoja}' declarada no existe en la plantilla."
        ) from error

    # Design's Sequence, step 4: validation precedes every mutation, so no
    # partial write is ever observable, even in-memory.
    _validar_completitud(estructura, valores)

    # Design's Sequence, step 5: swap the logo (if any) before writing
    # values, reusing the original image's anchor object (D4).
    _intercambiar_logo(hoja, tipo.logo)

    # Design's Sequence, step 6: write every present value into its anchor
    # cell.
    _escribir_valores(hoja, estructura, valores)

    # Backlog #11 (design D5): embed stored attachments into their declared
    # anchor slots, alongside the logo swap, before exporting.
    _incrustar_adjuntos(hoja, estructura, adjuntos)

    # Design's Sequence, step 7: export only the declared sheet.
    _exportar_solo_hoja_declarada(libro, nombre_hoja)

    # Design's Sequence, step 8: materialize the final bytes.
    buffer = BytesIO()
    libro.save(buffer)
    buffer.seek(0)
    return buffer

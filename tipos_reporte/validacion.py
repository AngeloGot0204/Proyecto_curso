"""Validation rules for a `DefinicionDeTipo.estructura` tree (spec:
"Exhaustive activation validation", "Closed data-type catalog", "Definition
names its sheet"; design D4, D5, D6).

R1-R4 (Slice 2) are pure functions over a plain dict — no database access,
no filesystem access, no `TipoDeReporte`/`DefinicionDeTipo` instance. R5-R6
(Slice 3) additionally take an OPEN BINARY FILE OBJECT for the `.xlsx`
template — never a `FieldFile` — so they stay testable against a real
in-memory workbook built by `openpyxl`, with no saved model required
(design D5). `validar_definicion` composes both groups into one
`ResultadoDeValidacion` and never returns early, so a document that fails
both groups reports both (settled decision 4: accumulate every problem).

`analizar_yaml_seguro` is the single deserialization entry point for
administrator-uploaded YAML (Threat Matrix: "Untrusted deserialization").
It is a thin, deliberate wrapper: `yaml.safe_load`, never `yaml.load` with
the default loader, because the latter constructs arbitrary Python objects
from attacker-controlled input.
"""

import json
from dataclasses import dataclass

import yaml
from django.core.exceptions import ValidationError
from openpyxl import load_workbook
from openpyxl.utils.cell import coordinate_from_string
from openpyxl.utils.exceptions import CellCoordinatesException
from yaml import YAMLError

from tipos_reporte.models import TipoDeDato

# Data types whose destination is a cell RANGE (celda_inicio + celda_fin)
# rather than a single celda (design's Interfaces/Contracts catalog table).
_TIPOS_CON_RANGO = {TipoDeDato.RANGO_HORA_INICIO_FIN}

# Characters that make coordinate_from_string tolerate a shape the
# generator does not support: absolute references, ranges, and
# sheet-qualified references (design's Interfaces/Contracts note).
_CARACTERES_NO_PERMITIDOS_EN_CELDA = ("$", ":", "!")


@dataclass(frozen=True)
class ProblemaDeDefinicion:
    """One accumulated validation failure.

    `regla` is a stable identifier assertable in tests (design D5) — it
    must never change when only the human-readable `mensaje` is reworded.
    """

    regla: str
    ubicacion: str
    mensaje: str


@dataclass(frozen=True)
class ResultadoDeValidacion:
    """The outcome of a full activation validation (design D5): every
    accumulated problem across R1-R6, plus a convenience `es_valida`
    property. Frozen, like `ProblemaDeDefinicion` — an outcome that already
    happened must not be mutated after the fact."""

    problemas: tuple[ProblemaDeDefinicion, ...]

    @property
    def es_valida(self) -> bool:
        return not self.problemas


def _es_celda_valida(valor) -> bool:
    if not isinstance(valor, str) or not valor:
        return False
    if any(caracter in valor for caracter in _CARACTERES_NO_PERMITIDOS_EN_CELDA):
        return False
    try:
        coordinate_from_string(valor)
    except (CellCoordinatesException, ValueError):
        return False
    return True


def _claves_de_celda_requeridas(tipo) -> list[str]:
    if tipo in _TIPOS_CON_RANGO:
        return ["celda_inicio", "celda_fin"]
    return ["celda"]


def _iterar_nodos(estructura: dict):
    """Yield `(ubicacion, nodo, clave_de_etiqueta)` for every campo/item in
    the tree. `clave_de_etiqueta` differs by node type: campos use
    `etiqueta`, items use `texto` (design's Interfaces/Contracts example)."""
    secciones = estructura.get("secciones") or []
    for i, seccion in enumerate(secciones):
        base = f"secciones[{i}]"
        for j, campo in enumerate(seccion.get("campos") or []):
            yield f"{base}.campos[{j}]", campo, "etiqueta"
        for j, item in enumerate(seccion.get("items") or []):
            yield f"{base}.items[{j}]", item, "texto"


def _validar_campos_obligatorios(ubicacion, nodo, clave_de_etiqueta):
    problemas = []

    def _requerir(clave):
        if not nodo.get(clave):
            problemas.append(
                ProblemaDeDefinicion(
                    regla="campo-obligatorio-ausente",
                    ubicacion=ubicacion,
                    mensaje=f"Falta la clave obligatoria '{clave}'.",
                )
            )

    _requerir("id")
    _requerir(clave_de_etiqueta)
    _requerir("tipo")

    tipo = nodo.get("tipo")
    if tipo in TipoDeDato.values:
        for clave in _claves_de_celda_requeridas(tipo):
            _requerir(clave)
        if tipo == TipoDeDato.SELECCION:
            _requerir("opciones")

    return problemas


def _validar_tipo_conocido(ubicacion, nodo):
    tipo = nodo.get("tipo")
    if tipo is not None and tipo not in TipoDeDato.values:
        return [
            ProblemaDeDefinicion(
                regla="tipo-de-dato-desconocido",
                ubicacion=ubicacion,
                mensaje=f"El tipo '{tipo}' no pertenece al catálogo cerrado.",
            )
        ]
    return []


def _validar_notacion_de_celda(ubicacion, nodo):
    problemas = []
    for clave in ("celda", "celda_inicio", "celda_fin"):
        valor = nodo.get(clave)
        if valor is None:
            continue
        if not _es_celda_valida(valor):
            problemas.append(
                ProblemaDeDefinicion(
                    regla="celda-mal-formada",
                    ubicacion=ubicacion,
                    mensaje=f"'{clave}': '{valor}' no es una notación de celda válida.",
                )
            )
    return problemas


def _validar_colisiones_de_celda(nodos):
    """Collision detection only considers already-valid cells (mirrors R6's
    documented precedent of only checking what already passed notation
    validation) — a shared typo is reported as `celda-mal-formada` for each
    occurrence, not additionally as a collision."""
    ocupantes: dict[str, list[str]] = {}
    for ubicacion, nodo, _clave_de_etiqueta in nodos:
        for clave in ("celda", "celda_inicio", "celda_fin"):
            valor = nodo.get(clave)
            if valor and _es_celda_valida(valor):
                ocupantes.setdefault(valor, []).append(ubicacion)

    problemas = []
    for celda, ubicaciones in ocupantes.items():
        if len(ubicaciones) > 1:
            for ubicacion in ubicaciones:
                problemas.append(
                    ProblemaDeDefinicion(
                        regla="celda-duplicada",
                        ubicacion=ubicacion,
                        mensaje=(
                            f"La celda '{celda}' está usada por más de un "
                            f"campo/ítem: {', '.join(ubicaciones)}."
                        ),
                    )
                )
    return problemas


_MAXIMO_DE_ANCLAS_DE_ADJUNTOS = 4


def _validar_anclas_de_adjuntos(estructura: dict) -> list[ProblemaDeDefinicion]:
    """R7 (backlog #11, design D6): validates the notation of every
    declared attachment anchor slot (`estructura["adjuntos"][*]["celda"]`,
    reusing `_es_celda_valida`, same as R3) and rejects more than
    `_MAXIMO_DE_ANCLAS_DE_ADJUNTOS` declared slots.

    Deliberately does NOT reuse R6's merged-anchor-cell rule
    (`celda-no-es-ancla`) or `_validar_colisiones_de_celda`: a floating
    image anchors to a cell CORNER, it is never written into the cell, so a
    merged non-anchor cell — or a cell shared with a data field — is a
    legitimate embedding target, not a collision (design D6)."""
    adjuntos = estructura.get("adjuntos") or []
    problemas = []
    for i, slot in enumerate(adjuntos):
        celda = slot.get("celda") if isinstance(slot, dict) else None
        if not _es_celda_valida(celda):
            problemas.append(
                ProblemaDeDefinicion(
                    regla="ancla-de-adjunto-mal-formada",
                    ubicacion=f"adjuntos[{i}]",
                    mensaje=f"'celda': '{celda}' no es una notación de celda válida.",
                )
            )
    if len(adjuntos) > _MAXIMO_DE_ANCLAS_DE_ADJUNTOS:
        problemas.append(
            ProblemaDeDefinicion(
                regla="anclas-de-adjunto-excedidas",
                ubicacion="adjuntos",
                mensaje=(
                    f"Se declararon {len(adjuntos)} anclas de adjuntos; el "
                    f"máximo permitido es {_MAXIMO_DE_ANCLAS_DE_ADJUNTOS}."
                ),
            )
        )
    return problemas


def validar_estructura(estructura: dict) -> list[ProblemaDeDefinicion]:
    """Runs R1-R4 and R7 over the full definition and accumulates every
    problem found (spec: "Exhaustive activation validation") — never
    returns early."""
    problemas: list[ProblemaDeDefinicion] = []
    nodos = list(_iterar_nodos(estructura))

    for ubicacion, nodo, clave_de_etiqueta in nodos:
        problemas.extend(_validar_campos_obligatorios(ubicacion, nodo, clave_de_etiqueta))
        problemas.extend(_validar_tipo_conocido(ubicacion, nodo))
        problemas.extend(_validar_notacion_de_celda(ubicacion, nodo))

    problemas.extend(_validar_colisiones_de_celda(nodos))
    problemas.extend(_validar_anclas_de_adjuntos(estructura))

    return problemas


def _mapa_de_celdas_no_ancla(hoja) -> dict[str, str]:
    """Maps every non-anchor cell of every merged range on `hoja` to the
    coordinate of its range's anchor (top-left) cell. A cell absent from
    this mapping either belongs to no merged range at all, or is itself an
    anchor — both are valid destinations (design's data-flow note:
    `read_only=True` must NOT be used, it does not populate `merged_cells`;
    `load_workbook`'s default already avoids that)."""
    mapa: dict[str, str] = {}
    for rango in hoja.merged_cells.ranges:
        ancla = rango.coord.split(":")[0]
        for fila in hoja[rango.coord]:
            for celda in fila:
                if celda.coordinate != ancla:
                    mapa[celda.coordinate] = ancla
    return mapa


def validar_contra_plantilla(estructura: dict, plantilla) -> list[ProblemaDeDefinicion]:
    """Runs R5 (template + declared sheet must be readable) and R6 (every
    destination cell must be the anchor of its merged range) against a REAL
    `.xlsx` workbook (spec: "Definition names its sheet", "Exhaustive
    activation validation"; design D5, D6, Threat Matrix).

    `plantilla` is an open binary file object, never a `FieldFile` — the
    caller is responsible for opening/closing it (design D5).

    If the template cannot be opened at all, this returns EXACTLY ONE
    problem (`plantilla-ilegible`) and skips every other check in this
    function — any exception from openpyxl becomes that one problem, never
    an uncaught exception (Threat Matrix: "Untrusted file parsing").
    """
    try:
        libro = load_workbook(plantilla)
    except Exception:
        return [
            ProblemaDeDefinicion(
                regla="plantilla-ilegible",
                ubicacion="plantilla",
                mensaje="No se pudo leer el archivo de plantilla (.xlsx).",
            )
        ]

    problemas: list[ProblemaDeDefinicion] = []
    nombre_hoja = estructura.get("hoja")
    if not nombre_hoja:
        problemas.append(
            ProblemaDeDefinicion(
                regla="hoja-ausente",
                ubicacion="hoja",
                mensaje="La definición debe declarar la clave 'hoja'.",
            )
        )
        return problemas

    if nombre_hoja not in libro.sheetnames:
        problemas.append(
            ProblemaDeDefinicion(
                regla="hoja-no-encontrada",
                ubicacion="hoja",
                mensaje=(
                    f"La hoja '{nombre_hoja}' declarada no existe en la "
                    "plantilla."
                ),
            )
        )
        return problemas

    hoja = libro[nombre_hoja]
    no_anclas = _mapa_de_celdas_no_ancla(hoja)

    # R6 only considers cells that already passed R3 (notation) — feeding a
    # typo to openpyxl would produce a second, derived complaint about the
    # same problem (design's documented precedent for this rule).
    for ubicacion, nodo, _clave_de_etiqueta in _iterar_nodos(estructura):
        for clave in ("celda", "celda_inicio", "celda_fin"):
            valor = nodo.get(clave)
            if valor and _es_celda_valida(valor) and valor in no_anclas:
                ancla = no_anclas[valor]
                problemas.append(
                    ProblemaDeDefinicion(
                        regla="celda-no-es-ancla",
                        ubicacion=ubicacion,
                        mensaje=(
                            f"'{clave}': '{valor}' no es la celda ancla de su "
                            f"rango combinado; usá '{ancla}'."
                        ),
                    )
                )

    return problemas


def validar_definicion(estructura: dict, plantilla=None) -> ResultadoDeValidacion:
    """Composes `validar_estructura` (R1-R4) and `validar_contra_plantilla`
    (R5-R6) into one `ResultadoDeValidacion`. Never returns early: both
    groups always run, so a document that fails both reports both (settled
    decision 4; design D5). `plantilla=None` skips R5-R6 — useful while a
    draft is being edited and no activation is being attempted."""
    problemas = list(validar_estructura(estructura))
    if plantilla is not None:
        problemas.extend(validar_contra_plantilla(estructura, plantilla))
    return ResultadoDeValidacion(problemas=tuple(problemas))


def analizar_yaml_seguro(texto: str):
    """The only deserialization entry point allowed for administrator-
    uploaded YAML (design D4, Threat Matrix: "Untrusted deserialization").
    A thin, deliberate wrapper around `yaml.safe_load` — never `yaml.load`
    with the default loader, which constructs arbitrary Python objects from
    attacker-controlled input (e.g. a `!!python/object/apply` tag)."""
    return yaml.safe_load(texto)


def analizar_definicion_subida(archivo) -> tuple[str, dict]:
    """UTF-8 decode → `analizar_yaml_seguro` → dict-root check →
    JSON-representability check, over an uploaded/stored definition file
    (backlog #13, S-14; design D2). Returns `(yaml_fuente, estructura)`.
    Raises Django's `ValidationError` keyed on `"archivo_yaml"` — the
    `DefinicionDeTipo` MODEL field name, so the key is domain-level, not
    form-level. Body moved verbatim from `admin.py::DefinicionDeTipoForm.
    clean()` (same four checks, same four Spanish messages), minus the
    `archivo is None` guard and the `cleaned[...]` assignments, which stay
    at each call site. Raising `ValidationError` rather than returning
    `ProblemaDeDefinicion`s is deliberate: this is a *fail-fast upload gate*
    ("can this become a JSON document?"), not the accumulating activation
    gate (R1-R7) `validar_definicion` runs above."""
    archivo.seek(0)
    try:
        texto = archivo.read().decode("utf-8")
    except UnicodeDecodeError as error:
        raise ValidationError(
            {"archivo_yaml": f"El archivo no es texto UTF-8 válido: {error}"}
        )
    archivo.seek(0)

    try:
        estructura = analizar_yaml_seguro(texto)
    except YAMLError as error:
        raise ValidationError({"archivo_yaml": f"YAML inválido: {error}"})

    if not isinstance(estructura, dict):
        raise ValidationError(
            {"archivo_yaml": "El documento debe ser un mapeo en su raíz."}
        )
    try:
        json.dumps(estructura)
    except (TypeError, ValueError):
        raise ValidationError(
            {
                "archivo_yaml": (
                    "El documento no es representable como JSON "
                    "(por ejemplo, contiene una fecha nativa de YAML)."
                )
            }
        )

    return texto, estructura

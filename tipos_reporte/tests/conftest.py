import copy

import openpyxl
import pytest
from django.core.files.uploadedfile import SimpleUploadedFile


@pytest.fixture
def definicion_valida():
    """A minimal but complete `estructura` dict: one section with one campo
    and one item, both fully specified, no collisions, all cells valid
    (design's Interfaces/Contracts example). Passes R1-R4 (Slice 2)
    unconditionally; `hoja`/template-anchor checks (R5-R6) are validated
    against it starting Slice 3, once `plantilla_xlsx` exists.

    Returns a factory (not the dict itself) so each test gets its own deep
    copy — callers mutate it freely (e.g. `del`, key reassignment) without
    leaking state into other tests."""

    def _crear():
        return copy.deepcopy(
            {
                "tipo": "instalacion-resinas",
                "plantilla": "JME.PC-0001.F1.xlsx",
                "hoja": "REPORTE",
                "secciones": [
                    {
                        "id": "datos-generales",
                        "titulo": "Datos generales",
                        "campos": [
                            {
                                "id": "turno",
                                "etiqueta": "Turno",
                                "tipo": "seleccion",
                                "opciones": ["Día", "Noche"],
                                "obligatorio": True,
                                "celda": "M12",
                            }
                        ],
                    },
                    {
                        "id": "proceso-instalacion",
                        "titulo": "Proceso de instalación",
                        "roles": ["construccion-jme", "qa-subterra"],
                        "items": [
                            {
                                "id": "p-01",
                                "texto": "Se verifica ángulo de perforación.",
                                "tipo": "rango-hora-inicio-fin",
                                "celda_inicio": "M25",
                                "celda_fin": "P25",
                            }
                        ],
                    },
                ],
            }
        )

    return _crear


@pytest.fixture
def imagen_png(tmp_path):
    """Distinct, identifiable PNGs: size is the discriminator after round
    trip (design's Test Fixture Extension). Returns a factory so a test can
    build several differently-sized/colored images (e.g. one for the
    template's original image, a distinct one for the tipo's logo)."""
    from PIL import Image

    def _crear(nombre="img.png", tamano=(10, 10), color=(255, 0, 0)):
        ruta = tmp_path / nombre
        Image.new("RGB", tamano, color).save(ruta)
        return ruta

    return _crear


@pytest.fixture
def plantilla_xlsx(tmp_path):
    """A real `.xlsx` workbook built with openpyxl, not committed to the
    repository (design's Testing Strategy). By default it declares one
    sheet named "REPORTE" with one merged range `M12:P12`, whose anchor
    (top-left) cell is `M12` — the exact shape ADR-0002 validated
    empirically. Returns a factory so each test can vary the sheet name,
    merged ranges, extra (non-target) sheets (`hojas_extra`, design's Test
    Fixture Extension — proves sheet-only export) and an embedded image
    (`imagen`, a path from `imagen_png` — proves the logo-swap scenarios).
    Both new kwargs default to today's behavior, so existing Slice-3 tests
    stay untouched."""

    def _crear(nombre_hoja="REPORTE", rangos=("M12:P12",), hojas_extra=(), imagen=None):
        wb = openpyxl.Workbook()
        wb.active.title = nombre_hoja
        for rango in rangos:
            wb.active.merge_cells(rango)
        for nombre in hojas_extra:
            wb.create_sheet(nombre)
        if imagen is not None:
            wb.active.add_image(openpyxl.drawing.image.Image(str(imagen)), "B2")
        destino = tmp_path / "plantilla.xlsx"
        wb.save(destino)
        return destino

    return _crear


@pytest.fixture
def valores_completos(definicion_valida):
    """A `valores` dict complete for `definicion_valida`'s default
    structure — one scalar `campo` (`turno`) and one range `item`
    (`p-01_inicio`/`p-01_fin`) — reusable across `test_generador.py`
    scenarios that need a passing completeness check (design's Testing
    Strategy). Returns a factory so callers can `dict(valores_completos())`
    and delete/override keys to build partial scenarios."""

    def _crear():
        return {
            "turno": "Día",
            "p-01_inicio": "08:00",
            "p-01_fin": "08:30",
        }

    return _crear


@pytest.fixture
def usuario_factory(db):
    """Create a Usuario with sensible defaults, overridable by kwargs.
    Mirrors `usuarios/tests/conftest.py`'s fixture (app-local test
    convention) — this app's admin tests (Slice 4) need one too and cross-
    app fixture sharing is not this project's pattern."""
    from usuarios.models import Usuario

    def _create(**kwargs):
        defaults = {
            "username": "usuario_test",
            "password": "irrelevant-not-hashed-for-this-fixture",
        }
        defaults.update(kwargs)
        return Usuario.objects.create(**defaults)

    return _create


@pytest.fixture
def tipo_de_reporte_factory(db):
    """Create a TipoDeReporte with sensible defaults, overridable by kwargs.

    `plantilla` is a required FileField (spec: "TipoDeReporte model"), so the
    factory always supplies a minimal in-memory file unless the caller
    overrides it.
    """
    from tipos_reporte.models import TipoDeReporte

    def _create(**kwargs):
        defaults = {
            "nombre": "Instalación de resinas",
            "codigo": "instalacion-resinas",
            "plantilla": SimpleUploadedFile(
                "plantilla.xlsx", b"contenido-irrelevante-para-este-nivel"
            ),
        }
        defaults.update(kwargs)
        return TipoDeReporte.objects.create(**defaults)

    return _create


@pytest.fixture
def administrador_factory(db):
    """Create a Usuario with `rol=Rol.ADMINISTRADOR` (backlog #13, S-14;
    design D1's `solo_administradores` gate), overridable by kwargs. App-
    local convention, mirrors `usuario_factory` above/`usuarios/tests/
    conftest.py`."""
    from usuarios.models import Rol, Usuario

    def _create(**kwargs):
        defaults = {
            "username": "administrador_test",
            "password": "irrelevant-not-hashed-for-this-fixture",
            "rol": Rol.ADMINISTRADOR,
        }
        defaults.update(kwargs)
        return Usuario.objects.create(**defaults)

    return _create


@pytest.fixture
def definicion_factory(db):
    """Create a `DefinicionDeTipo` with sensible defaults, overridable by
    kwargs (backlog #13, S-14). Mirrors `test_activacion.py`'s local
    `_crear_borrador` helper. `tipo` is required — the caller must pass one
    (e.g. from `tipo_de_reporte_factory`)."""
    from tipos_reporte.models import DefinicionDeTipo, Estado

    def _create(**kwargs):
        defaults = {
            "archivo_yaml": SimpleUploadedFile(
                "definicion.yaml", b"secciones: []"
            ),
            "yaml_fuente": "secciones: []",
            "estructura": {"secciones": []},
            "estado": Estado.BORRADOR,
        }
        defaults.update(kwargs)
        return DefinicionDeTipo.objects.create(**defaults)

    return _create

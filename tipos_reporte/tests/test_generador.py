"""Tests for `tipos_reporte.generador.generar_reporte` (backlog #4).

Strict TDD: every scenario below is written RED (failing, referencing
production code that does not exist yet or new behavior it does not yet
implement) before the corresponding implementation lands in
`tipos_reporte/generador.py`. See
`openspec/changes/generador-excel-plantilla/specs/generacion-reporte-excel/spec.md`
for the acceptance criteria these tests encode, and `design.md` for the
architecture decisions (D1-D5) they exercise.

This file covers Phase 1 (exceptions), Phase 2 (template loading +
completeness validation), Phase 3 (cell writing + sheet-only export) and
Phase 4 (logo swap) of the change's `tasks.md`.
"""

import os
from io import BytesIO

import pytest
from django.core.files.uploadedfile import SimpleUploadedFile
from openpyxl import load_workbook
from PIL import Image

from tipos_reporte.models import DefinicionDeTipo, Estado


def _definicion_con_plantilla(
    tipo_de_reporte_factory, plantilla_xlsx, estructura, **kwargs_tipo
):
    """Build a `DefinicionDeTipo` whose `tipo.plantilla` is a real `.xlsx`
    matching `estructura["hoja"]` (mirrors `test_activacion.py`'s
    `_tipo_con_plantilla` pattern). `generar_reporte` does not check
    `estado` itself (design's Sequence — it trusts an already-activated
    definition), so a plain `borrador` row is enough for these tests."""
    destino = plantilla_xlsx(nombre_hoja=estructura["hoja"])
    with open(destino, "rb") as archivo:
        contenido = archivo.read()
    defaults = {"plantilla": SimpleUploadedFile("plantilla.xlsx", contenido)}
    defaults.update(kwargs_tipo)
    tipo = tipo_de_reporte_factory(**defaults)
    return DefinicionDeTipo.objects.create(
        tipo=tipo,
        archivo_yaml=SimpleUploadedFile("definicion.yaml", b"secciones: []"),
        yaml_fuente="secciones: []",
        estructura=estructura,
        estado=Estado.BORRADOR,
    )


def test_problema_de_generacion_es_importable_y_es_excepcion():
    """Task 1.1 RED: `ProblemaDeGeneracion` must exist and be an `Exception`
    subclass, mirroring `validacion.ProblemaDeDefinicion`'s role as the
    stable base every foreseeable generation failure inherits from."""
    from tipos_reporte.generador import ProblemaDeGeneracion

    assert issubclass(ProblemaDeGeneracion, Exception)


def test_valores_incompletos_ordena_faltantes_y_los_incluye_en_el_mensaje():
    """Task 1.3 RED: `ValoresIncompletos(["b", "a"])` must sort `.faltantes`
    (design's Interfaces/Contracts) and mention every missing id in the
    exception message, so a caller can log or display it directly."""
    from tipos_reporte.generador import ValoresIncompletos

    excepcion = ValoresIncompletos(["b", "a"])

    assert excepcion.faltantes == ("a", "b")
    mensaje = str(excepcion)
    assert "a" in mensaje
    assert "b" in mensaje


@pytest.mark.django_db
def test_plantilla_carga_correctamente(
    tipo_de_reporte_factory, plantilla_xlsx, definicion_valida, valores_completos
):
    """Task 2.1 RED — spec scenario "Template loads successfully": a valid
    `plantilla_xlsx` on an activated-shaped `DefinicionDeTipo`, with
    complete `valores`, must return a `BytesIO` re-openable via
    `load_workbook`."""
    from tipos_reporte.generador import generar_reporte

    definicion = _definicion_con_plantilla(
        tipo_de_reporte_factory, plantilla_xlsx, definicion_valida()
    )

    resultado = generar_reporte(definicion, valores_completos())

    assert isinstance(resultado, BytesIO)
    libro = load_workbook(resultado)
    assert "REPORTE" in libro.sheetnames


@pytest.mark.django_db
def test_plantilla_con_bytes_invalidos_lanza_plantilla_ilegible(
    tipo_de_reporte_factory, definicion_valida, valores_completos
):
    """Task 2.2 RED — spec scenario "Template file cannot be read" (case 1):
    the factory's default blob is not a real workbook, so `generar_reporte`
    must raise `PlantillaIlegible`, never a raw `openpyxl`/IO exception."""
    from tipos_reporte.generador import PlantillaIlegible, generar_reporte

    tipo = tipo_de_reporte_factory()
    definicion = DefinicionDeTipo.objects.create(
        tipo=tipo,
        archivo_yaml=SimpleUploadedFile("definicion.yaml", b"secciones: []"),
        yaml_fuente="secciones: []",
        estructura=definicion_valida(),
        estado=Estado.BORRADOR,
    )

    with pytest.raises(PlantillaIlegible):
        generar_reporte(definicion, valores_completos())


def _estructura_completitud():
    """A structure with one required scalar (`supervisor`), one required
    range (`descanso`) and one OPTIONAL range (`opcional`, `obligatorio`
    absent) — reused across the completeness-validation scenarios (spec's
    "Missing Required Values" requirement; design D3)."""
    return {
        "tipo": "instalacion-resinas",
        "plantilla": "JME.PC-0001.F1.xlsx",
        "hoja": "REPORTE",
        "secciones": [
            {
                "id": "datos-generales",
                "titulo": "Datos generales",
                "campos": [
                    {
                        "id": "supervisor",
                        "etiqueta": "Supervisor",
                        "tipo": "texto",
                        "obligatorio": True,
                        "celda": "B2",
                    }
                ],
                "items": [
                    {
                        "id": "descanso",
                        "texto": "Descanso",
                        "tipo": "rango-hora-inicio-fin",
                        "obligatorio": True,
                        "celda_inicio": "C3",
                        "celda_fin": "C4",
                    },
                    {
                        "id": "opcional",
                        "texto": "Ítem opcional",
                        "tipo": "rango-hora-inicio-fin",
                        "celda_inicio": "D5",
                        "celda_fin": "D6",
                    },
                ],
            }
        ],
    }


def test_destinos_de_un_campo_escalar_devuelve_un_par_id_celda():
    """Task 2.4 RED (case 1): a scalar `campo` node must map to exactly one
    `(id, celda)` pair — design D1's single derivation of value key ->
    target cell."""
    from tipos_reporte.generador import _destinos

    nodo = {"id": "turno", "tipo": "texto", "celda": "B2"}

    assert _destinos(nodo) == [("turno", "B2")]


def test_destinos_de_un_rango_devuelve_dos_pares_inicio_fin():
    """Task 2.4 RED (case 2): a `rango-hora-inicio-fin` node must map to two
    independent `(id + sufijo, celda)` pairs (proposal's Values-Dict
    Contract — inicio/fin are two rows, not one composite value)."""
    from tipos_reporte.generador import _destinos

    nodo = {
        "id": "descanso",
        "tipo": "rango-hora-inicio-fin",
        "celda_inicio": "C3",
        "celda_fin": "C4",
    }

    assert _destinos(nodo) == [
        ("descanso_inicio", "C3"),
        ("descanso_fin", "C4"),
    ]


def test_claves_de_valor_de_un_campo_escalar_coincide_con_destinos():
    """PR 1 task 1.1 RED — design D5: `claves_de_valor(nodo)` is the public
    extraction of `_destinos`' key derivation. For a scalar node it must
    return exactly the same key `_destinos` currently pairs with a
    coordinate — one-element list, no coordinate."""
    from tipos_reporte.generador import _destinos, claves_de_valor

    nodo = {"id": "turno", "tipo": "texto", "celda": "B2"}

    assert claves_de_valor(nodo) == [clave for clave, _coordenada in _destinos(nodo)]
    assert claves_de_valor(nodo) == ["turno"]


def test_claves_de_valor_de_un_rango_coincide_con_destinos():
    """PR 1 task 1.1 RED — design D5, range case: `claves_de_valor` must
    derive the same two `_inicio`/`_fin` keys `_destinos` currently
    derives for a `rango-hora-inicio-fin` node."""
    from tipos_reporte.generador import _destinos, claves_de_valor

    nodo = {
        "id": "descanso",
        "tipo": "rango-hora-inicio-fin",
        "celda_inicio": "C3",
        "celda_fin": "C4",
    }

    assert claves_de_valor(nodo) == [clave for clave, _coordenada in _destinos(nodo)]
    assert claves_de_valor(nodo) == ["descanso_inicio", "descanso_fin"]


@pytest.mark.django_db
def test_plantilla_faltante_en_storage_lanza_plantilla_ilegible(
    tipo_de_reporte_factory, plantilla_xlsx, definicion_valida, valores_completos
):
    """Task 2.2 RED — spec scenario "Template file cannot be read" (case 2):
    a deleted/missing file in storage must also raise `PlantillaIlegible`
    (mirrors `activar_definicion`'s own guard for `plantilla.open("rb")`)."""
    from tipos_reporte.generador import PlantillaIlegible, generar_reporte

    definicion = _definicion_con_plantilla(
        tipo_de_reporte_factory, plantilla_xlsx, definicion_valida()
    )
    os.remove(definicion.tipo.plantilla.path)

    with pytest.raises(PlantillaIlegible):
        generar_reporte(definicion, valores_completos())


@pytest.mark.django_db
def test_falta_un_valor_simple_requerido_lanza_valores_incompletos(
    tipo_de_reporte_factory, plantilla_xlsx
):
    """Task 2.6 RED — spec scenario "A required simple value is missing":
    a required `campo` (`supervisor`) absent from `valores` must raise
    `ValoresIncompletos` listing it, and return no bytes."""
    from tipos_reporte.generador import ValoresIncompletos, generar_reporte

    estructura = _estructura_completitud()
    definicion = _definicion_con_plantilla(tipo_de_reporte_factory, plantilla_xlsx, estructura)
    valores = {"descanso_inicio": "08:00", "descanso_fin": "08:30"}

    with pytest.raises(ValoresIncompletos) as info:
        generar_reporte(definicion, valores)

    assert "supervisor" in info.value.faltantes


@pytest.mark.django_db
def test_falta_un_lado_de_un_rango_requerido_lanza_valores_incompletos(
    tipo_de_reporte_factory, plantilla_xlsx
):
    """Task 2.7 RED — spec scenario "Only one side of a required range
    value is missing": `descanso_inicio` present, `descanso_fin` absent
    must raise `ValoresIncompletos` with exactly `("descanso_fin",)`."""
    from tipos_reporte.generador import ValoresIncompletos, generar_reporte

    estructura = _estructura_completitud()
    definicion = _definicion_con_plantilla(tipo_de_reporte_factory, plantilla_xlsx, estructura)
    valores = {"supervisor": "Ana", "descanso_inicio": "08:00"}

    with pytest.raises(ValoresIncompletos) as info:
        generar_reporte(definicion, valores)

    assert info.value.faltantes == ("descanso_fin",)


@pytest.mark.django_db
def test_multiples_ids_faltantes_se_reportan_juntos(
    tipo_de_reporte_factory, plantilla_xlsx
):
    """Task 2.8 RED — spec scenario "Multiple missing ids are all reported
    together": two required ids missing must both appear in ONE raised
    `.faltantes`, not fail-fast on the first (design's "accumulate every
    problem")."""
    from tipos_reporte.generador import ValoresIncompletos, generar_reporte

    estructura = _estructura_completitud()
    definicion = _definicion_con_plantilla(tipo_de_reporte_factory, plantilla_xlsx, estructura)
    valores = {}

    with pytest.raises(ValoresIncompletos) as info:
        generar_reporte(definicion, valores)

    assert "supervisor" in info.value.faltantes
    assert "descanso_inicio" in info.value.faltantes
    assert "descanso_fin" in info.value.faltantes


@pytest.mark.django_db
def test_item_no_obligatorio_con_clave_ausente_no_lanza(
    tipo_de_reporte_factory, plantilla_xlsx
):
    """Task 2.9 RED — completeness confirms only required leaf ids raise:
    the `opcional` item (no `obligatorio` key, design D3) has BOTH its
    `_inicio`/`_fin` keys absent from `valores`, yet generation must
    succeed as long as every REQUIRED id is present."""
    from tipos_reporte.generador import generar_reporte

    estructura = _estructura_completitud()
    definicion = _definicion_con_plantilla(tipo_de_reporte_factory, plantilla_xlsx, estructura)
    valores = {
        "supervisor": "Ana",
        "descanso_inicio": "08:00",
        "descanso_fin": "08:30",
    }

    resultado = generar_reporte(definicion, valores)

    assert isinstance(resultado, BytesIO)


@pytest.mark.django_db
def test_valor_de_campo_simple_se_escribe_por_id(
    tipo_de_reporte_factory, plantilla_xlsx
):
    """Task 3.1 RED — spec scenario "Simple field value is written by id":
    a `campo` with `id="turno"`/`celda="B2"` and `valores={"turno":
    "Mañana"}` must leave `B2` containing `"Mañana"` in the exported
    sheet."""
    from tipos_reporte.generador import generar_reporte

    estructura = {
        "tipo": "instalacion-resinas",
        "plantilla": "JME.PC-0001.F1.xlsx",
        "hoja": "REPORTE",
        "secciones": [
            {
                "id": "datos-generales",
                "titulo": "Datos generales",
                "campos": [
                    {
                        "id": "turno",
                        "etiqueta": "Turno",
                        "tipo": "texto",
                        "obligatorio": True,
                        "celda": "B2",
                    }
                ],
            }
        ],
    }
    definicion = _definicion_con_plantilla(tipo_de_reporte_factory, plantilla_xlsx, estructura)

    resultado = generar_reporte(definicion, {"turno": "Mañana"})

    libro = load_workbook(resultado)
    hoja = libro["REPORTE"]
    assert hoja["B2"].value == "Mañana"


@pytest.mark.django_db
def test_valores_de_rango_se_escriben_desde_dos_claves_independientes(
    tipo_de_reporte_factory, plantilla_xlsx
):
    """Task 3.2 RED — spec scenario "Range field values are written from
    two independent keys": `descanso_inicio` -> `C3`, `descanso_fin` ->
    `C4`, written independently."""
    from tipos_reporte.generador import generar_reporte

    estructura = _estructura_completitud()
    definicion = _definicion_con_plantilla(tipo_de_reporte_factory, plantilla_xlsx, estructura)
    valores = {
        "supervisor": "Ana",
        "descanso_inicio": "08:00",
        "descanso_fin": "08:30",
    }

    resultado = generar_reporte(definicion, valores)

    libro = load_workbook(resultado)
    hoja = libro["REPORTE"]
    assert hoja["C3"].value == "08:00"
    assert hoja["C4"].value == "08:30"


@pytest.mark.django_db
@pytest.mark.parametrize("valor", [False, 0])
def test_valores_falsy_se_escriben_tal_cual(
    tipo_de_reporte_factory, plantilla_xlsx, valor
):
    """Task 3.3 RED — design D2: `False`/`0` are legitimate values and must
    be written as-is, never skipped as if absent (membership test, not
    truthiness). `""` is covered separately at the unit level
    (`test_escribir_valores_escribe_string_vacio_tal_cual`) because
    openpyxl/XLSX round-trips an empty-string cell as `None` — the same
    bytes a never-written cell would produce — so a save/reload assertion
    cannot distinguish "written empty string" from "left untouched"."""
    from tipos_reporte.generador import generar_reporte

    estructura = {
        "tipo": "instalacion-resinas",
        "plantilla": "JME.PC-0001.F1.xlsx",
        "hoja": "REPORTE",
        "secciones": [
            {
                "id": "datos-generales",
                "titulo": "Datos generales",
                "campos": [
                    {
                        "id": "activo",
                        "etiqueta": "Activo",
                        "tipo": "booleano",
                        "obligatorio": True,
                        "celda": "B2",
                    }
                ],
            }
        ],
    }
    definicion = _definicion_con_plantilla(tipo_de_reporte_factory, plantilla_xlsx, estructura)

    resultado = generar_reporte(definicion, {"activo": valor})

    libro = load_workbook(resultado)
    hoja = libro["REPORTE"]
    assert hoja["B2"].value == valor


def test_escribir_valores_escribe_string_vacio_tal_cual():
    """Task 3.3 RED (unit-level companion): `_escribir_valores` must assign
    `""` to the cell object (membership test, D2) even though XLSX itself
    cannot distinguish that from a never-written cell after a save/reload
    round trip — asserted directly on the in-memory `openpyxl` worksheet,
    before any serialization happens."""
    import openpyxl

    from tipos_reporte.generador import _escribir_valores

    estructura = {
        "hoja": "REPORTE",
        "secciones": [
            {
                "id": "datos-generales",
                "campos": [
                    {"id": "activo", "tipo": "booleano", "celda": "B2"}
                ],
            }
        ],
    }
    hoja = openpyxl.Workbook().active

    _escribir_valores(hoja, estructura, {"activo": ""})

    assert hoja["B2"].value == ""


@pytest.mark.django_db
def test_solo_se_exporta_la_hoja_declarada(
    tipo_de_reporte_factory, plantilla_xlsx, definicion_valida, valores_completos
):
    """Task 3.5 RED — spec scenario "Only the declared sheet is exported":
    a template with an extra sheet (`hojas_extra=("Otra",)`) must export
    only `estructura["hoja"]`."""
    from tipos_reporte.generador import generar_reporte

    estructura = definicion_valida()
    destino = plantilla_xlsx(nombre_hoja=estructura["hoja"], hojas_extra=("Otra",))
    with open(destino, "rb") as archivo:
        contenido = archivo.read()
    tipo = tipo_de_reporte_factory(plantilla=SimpleUploadedFile("plantilla.xlsx", contenido))
    definicion = DefinicionDeTipo.objects.create(
        tipo=tipo,
        archivo_yaml=SimpleUploadedFile("definicion.yaml", b"secciones: []"),
        yaml_fuente="secciones: []",
        estructura=estructura,
        estado=Estado.BORRADOR,
    )

    resultado = generar_reporte(definicion, valores_completos())

    libro = load_workbook(resultado)
    assert libro.sheetnames == [estructura["hoja"]]


@pytest.mark.django_db
def test_contenido_no_tocado_de_la_hoja_permanece_estructuralmente_identico(
    tipo_de_reporte_factory, plantilla_xlsx, definicion_valida, valores_completos
):
    """Task 3.6 RED — spec scenario "Untouched sheet content remains
    byte-identical in structure": merged ranges outside the anchor cells
    must survive generation unchanged."""
    from tipos_reporte.generador import generar_reporte

    estructura = definicion_valida()
    destino = plantilla_xlsx(
        nombre_hoja=estructura["hoja"], rangos=("M12:P12", "A1:B1")
    )
    with open(destino, "rb") as archivo:
        contenido = archivo.read()
    tipo = tipo_de_reporte_factory(plantilla=SimpleUploadedFile("plantilla.xlsx", contenido))
    definicion = DefinicionDeTipo.objects.create(
        tipo=tipo,
        archivo_yaml=SimpleUploadedFile("definicion.yaml", b"secciones: []"),
        yaml_fuente="secciones: []",
        estructura=estructura,
        estado=Estado.BORRADOR,
    )

    resultado = generar_reporte(definicion, valores_completos())

    libro = load_workbook(resultado)
    hoja = libro[estructura["hoja"]]
    rangos_resultantes = {str(rango) for rango in hoja.merged_cells.ranges}
    assert rangos_resultantes == {"M12:P12", "A1:B1"}


@pytest.mark.django_db
def test_generacion_exitosa_devuelve_bytes_legibles(
    tipo_de_reporte_factory, plantilla_xlsx, definicion_valida, valores_completos
):
    """Task 3.8 RED — spec scenario "Successful generation returns readable
    bytes": the final `BytesIO` must re-open via `load_workbook` without
    error, and its cursor must have been reset to the start."""
    from tipos_reporte.generador import generar_reporte

    definicion = _definicion_con_plantilla(
        tipo_de_reporte_factory, plantilla_xlsx, definicion_valida()
    )

    resultado = generar_reporte(definicion, valores_completos())

    assert resultado.tell() == 0
    libro = load_workbook(resultado)
    assert libro.sheetnames == ["REPORTE"]


@pytest.mark.django_db
def test_logo_presente_reemplaza_la_imagen_de_la_plantilla_en_el_mismo_anclaje(
    tipo_de_reporte_factory, plantilla_xlsx, imagen_png, definicion_valida, valores_completos
):
    """Task 4.1 RED — spec scenario "Logo is present on the tipo": a
    template with a 10x10 original image and a tipo whose `logo` is a
    distinct 20x20 PNG must export the tipo's logo — not the template's
    original image — at the original anchor (design D4: reuse the loaded
    `anchor` object, so position survives even though pixel size changes)."""
    from tipos_reporte.generador import generar_reporte

    estructura = definicion_valida()
    ruta_imagen_plantilla = imagen_png(nombre="original.png", tamano=(10, 10))
    destino = plantilla_xlsx(nombre_hoja=estructura["hoja"], imagen=ruta_imagen_plantilla)
    with open(destino, "rb") as archivo:
        contenido = archivo.read()
    ruta_logo = imagen_png(nombre="logo.png", tamano=(20, 20), color=(0, 255, 0))
    tipo = tipo_de_reporte_factory(
        plantilla=SimpleUploadedFile("plantilla.xlsx", contenido),
        logo=SimpleUploadedFile("logo.png", ruta_logo.read_bytes()),
    )
    definicion = DefinicionDeTipo.objects.create(
        tipo=tipo,
        archivo_yaml=SimpleUploadedFile("definicion.yaml", b"secciones: []"),
        yaml_fuente="secciones: []",
        estructura=estructura,
        estado=Estado.BORRADOR,
    )

    resultado = generar_reporte(definicion, valores_completos())

    libro = load_workbook(resultado)
    hoja = libro[estructura["hoja"]]
    assert len(hoja._images) == 1
    imagen_exportada = Image.open(hoja._images[0].ref)
    assert imagen_exportada.size == (20, 20)


@pytest.mark.django_db
def test_logo_ausente_deja_la_imagen_original_de_la_plantilla_intacta(
    tipo_de_reporte_factory, plantilla_xlsx, imagen_png, definicion_valida, valores_completos
):
    """Task 4.2 RED — spec scenario "Logo is absent on the tipo": with
    `tipo.logo` unset, the template's original 10x10 image must remain
    untouched — same size and same anchor position — after generation."""
    from tipos_reporte.generador import generar_reporte

    estructura = definicion_valida()
    ruta_imagen_plantilla = imagen_png(nombre="original.png", tamano=(10, 10))
    destino = plantilla_xlsx(nombre_hoja=estructura["hoja"], imagen=ruta_imagen_plantilla)
    with open(destino, "rb") as archivo:
        contenido = archivo.read()
    tipo = tipo_de_reporte_factory(plantilla=SimpleUploadedFile("plantilla.xlsx", contenido))
    definicion = DefinicionDeTipo.objects.create(
        tipo=tipo,
        archivo_yaml=SimpleUploadedFile("definicion.yaml", b"secciones: []"),
        yaml_fuente="secciones: []",
        estructura=estructura,
        estado=Estado.BORRADOR,
    )
    libro_plantilla = load_workbook(destino)
    ancla_original = libro_plantilla[estructura["hoja"]]._images[0].anchor

    resultado = generar_reporte(definicion, valores_completos())

    libro = load_workbook(resultado)
    hoja = libro[estructura["hoja"]]
    assert len(hoja._images) == 1
    imagen_exportada = Image.open(hoja._images[0].ref)
    assert imagen_exportada.size == (10, 10)
    ancla_exportada = hoja._images[0].anchor
    assert ancla_exportada._from.col == ancla_original._from.col
    assert ancla_exportada._from.row == ancla_original._from.row


@pytest.mark.django_db
def test_plantilla_sin_imagen_con_logo_definido_no_inserta_ninguna_imagen(
    tipo_de_reporte_factory, plantilla_xlsx, imagen_png, definicion_valida, valores_completos
):
    """Task 4.3 RED — "template has no image, logo is set": with no
    original image to anchor to, the logo must NOT be inserted at a
    hardcoded cell (design D4's rejected alternative) — `hoja._images`
    stays empty."""
    from tipos_reporte.generador import generar_reporte

    estructura = definicion_valida()
    destino = plantilla_xlsx(nombre_hoja=estructura["hoja"])
    with open(destino, "rb") as archivo:
        contenido = archivo.read()
    ruta_logo = imagen_png(nombre="logo.png", tamano=(20, 20), color=(0, 255, 0))
    tipo = tipo_de_reporte_factory(
        plantilla=SimpleUploadedFile("plantilla.xlsx", contenido),
        logo=SimpleUploadedFile("logo.png", ruta_logo.read_bytes()),
    )
    definicion = DefinicionDeTipo.objects.create(
        tipo=tipo,
        archivo_yaml=SimpleUploadedFile("definicion.yaml", b"secciones: []"),
        yaml_fuente="secciones: []",
        estructura=estructura,
        estado=Estado.BORRADOR,
    )

    resultado = generar_reporte(definicion, valores_completos())

    libro = load_workbook(resultado)
    hoja = libro[estructura["hoja"]]
    assert hoja._images == []


# --- Phase 5: attachment embedding via anchor slots (backlog #11, D5/D6) ---
#
# `generacion-reporte-excel` delta spec, "Attachment Embedding via Anchor
# Slots": `_incrustar_adjuntos` is a generalization of, not a reuse of,
# `_intercambiar_logo`'s single-fixed-anchor mechanism — it uses openpyxl's
# STRING anchor form (`hoja.add_image(img, "B40")`), so `Image.width/height`
# DO define the drawing box (the inverse of `_intercambiar_logo`, which
# copies a pre-existing anchor object and ignores `Image.width/height`).


def _cuatro_slots_de_adjuntos():
    return [
        {"celda": "B40"},
        {"celda": "B60"},
        {"celda": "B80"},
        {"celda": "B100"},
    ]


@pytest.mark.django_db
def test_adjuntos_dentro_del_limite_de_anclas_se_incrustan(
    tipo_de_reporte_factory, plantilla_xlsx, imagen_png, definicion_valida, valores_completos
):
    """Spec scenario "Attachments within anchor-slot count are embedded": 3
    stored attachments and a template declaring 4 anchor slots must embed
    all 3, each at its corresponding anchor slot."""
    from tipos_reporte.generador import generar_reporte

    estructura = definicion_valida()
    estructura["adjuntos"] = _cuatro_slots_de_adjuntos()
    destino = plantilla_xlsx(nombre_hoja=estructura["hoja"])
    with open(destino, "rb") as archivo:
        contenido = archivo.read()
    tipo = tipo_de_reporte_factory(plantilla=SimpleUploadedFile("plantilla.xlsx", contenido))
    definicion = DefinicionDeTipo.objects.create(
        tipo=tipo,
        archivo_yaml=SimpleUploadedFile("definicion.yaml", b"secciones: []"),
        yaml_fuente="secciones: []",
        estructura=estructura,
        estado=Estado.BORRADOR,
    )
    adjuntos = [
        open(imagen_png(nombre=f"adjunto-{i}.png", tamano=(400, 300)), "rb")
        for i in range(3)
    ]

    try:
        resultado = generar_reporte(definicion, valores_completos(), adjuntos=adjuntos)
    finally:
        for archivo in adjuntos:
            archivo.close()

    libro = load_workbook(resultado)
    hoja = libro[estructura["hoja"]]
    assert len(hoja._images) == 3
    anclas_esperadas = {(1, 39), (1, 59), (1, 79)}  # (col, row), 0-indexed
    anclas_obtenidas = {
        (imagen.anchor._from.col, imagen.anchor._from.row) for imagen in hoja._images
    }
    assert anclas_obtenidas == anclas_esperadas


@pytest.mark.django_db
def test_adjuntos_mas_alla_del_limite_de_anclas_quedan_almacenados_no_incrustados(
    tipo_de_reporte_factory, plantilla_xlsx, imagen_png, definicion_valida, valores_completos
):
    """Spec scenario "Attachments beyond anchor-slot count remain stored,
    not embedded": 6 stored attachments and 4 declared anchor slots must
    embed only 4 — `zip` truncation, no error raised."""
    from tipos_reporte.generador import generar_reporte

    estructura = definicion_valida()
    estructura["adjuntos"] = _cuatro_slots_de_adjuntos()
    destino = plantilla_xlsx(nombre_hoja=estructura["hoja"])
    with open(destino, "rb") as archivo:
        contenido = archivo.read()
    tipo = tipo_de_reporte_factory(plantilla=SimpleUploadedFile("plantilla.xlsx", contenido))
    definicion = DefinicionDeTipo.objects.create(
        tipo=tipo,
        archivo_yaml=SimpleUploadedFile("definicion.yaml", b"secciones: []"),
        yaml_fuente="secciones: []",
        estructura=estructura,
        estado=Estado.BORRADOR,
    )
    adjuntos = [
        open(imagen_png(nombre=f"adjunto-{i}.png", tamano=(400, 300)), "rb")
        for i in range(6)
    ]

    try:
        resultado = generar_reporte(definicion, valores_completos(), adjuntos=adjuntos)
    finally:
        for archivo in adjuntos:
            archivo.close()

    libro = load_workbook(resultado)
    hoja = libro[estructura["hoja"]]
    assert len(hoja._images) == 4


@pytest.mark.django_db
def test_sin_adjuntos_deja_las_anclas_vacias(
    tipo_de_reporte_factory, plantilla_xlsx, definicion_valida, valores_completos
):
    """Spec scenario "No attachments leaves anchor slots empty": a
    `Reporte` with no stored attachments and a template declaring anchor
    slots must leave the exported sheet's anchor slots empty, and
    generation must still succeed normally."""
    from tipos_reporte.generador import generar_reporte

    estructura = definicion_valida()
    estructura["adjuntos"] = _cuatro_slots_de_adjuntos()
    definicion = _definicion_con_plantilla(tipo_de_reporte_factory, plantilla_xlsx, estructura)

    resultado = generar_reporte(definicion, valores_completos(), adjuntos=())

    libro = load_workbook(resultado)
    hoja = libro[estructura["hoja"]]
    assert hoja._images == []


@pytest.mark.django_db
def test_plantilla_sin_anclas_declaradas_no_incrusta_nada(
    tipo_de_reporte_factory, plantilla_xlsx, imagen_png, definicion_valida, valores_completos
):
    """Spec scenario "Template without declared anchor slots skips
    embedding entirely": a `DefinicionDeTipo` whose template declares no
    `adjuntos` anchor slots must perform no embedding at all, even with
    stored attachments — generation proceeds as before this change (cell
    values and logo swap only)."""
    from tipos_reporte.generador import generar_reporte

    estructura = definicion_valida()  # no "adjuntos" key
    definicion = _definicion_con_plantilla(tipo_de_reporte_factory, plantilla_xlsx, estructura)
    adjunto = open(imagen_png(nombre="adjunto.png", tamano=(400, 300)), "rb")

    try:
        resultado = generar_reporte(
            definicion, valores_completos(), adjuntos=[adjunto]
        )
    finally:
        adjunto.close()

    libro = load_workbook(resultado)
    hoja = libro[estructura["hoja"]]
    assert hoja._images == []


def test_intercambiar_logo_logo_vacio_deja_imagen_de_plantilla_intacta():
    """8.1 RED (backlog #13, S-14; spec "Generation with no logo leaves the
    template default untouched"): a unit-level regression test, calling
    `_intercambiar_logo` directly with an empty/falsy `logo` — must leave
    `hoja._images` completely untouched (same list, same single original
    image), never removing or swapping anything (design D8)."""
    import openpyxl
    from openpyxl.drawing.image import Image as ImagenOpenpyxl

    from tipos_reporte.generador import _intercambiar_logo

    hoja = openpyxl.Workbook().active
    imagen_original = ImagenOpenpyxl.__new__(ImagenOpenpyxl)
    hoja._images = [imagen_original]

    _intercambiar_logo(hoja, None)

    assert hoja._images == [imagen_original]


@pytest.mark.django_db
def test_adjunto_no_decodificable_se_omite_sin_interrumpir_la_generacion(
    tipo_de_reporte_factory, plantilla_xlsx, definicion_valida, valores_completos
):
    """Threat Matrix "Untrusted file parsing": a stored file Pillow cannot
    decode (e.g. an unconverted HEIC that reached storage because client
    conversion failed) must be SKIPPED, never raised as
    `ProblemaDeGeneracion` — the isolation philosophy already established
    server-side in PR 2 ("bloqueo solo del adjunto"), one layer down."""
    from tipos_reporte.generador import generar_reporte

    estructura = definicion_valida()
    estructura["adjuntos"] = [{"celda": "B40"}]
    definicion = _definicion_con_plantilla(tipo_de_reporte_factory, plantilla_xlsx, estructura)
    adjunto_no_decodificable = SimpleUploadedFile(
        "adjunto.heic", b"esto no es una imagen decodificable"
    )

    resultado = generar_reporte(
        definicion, valores_completos(), adjuntos=[adjunto_no_decodificable]
    )

    libro = load_workbook(resultado)
    hoja = libro[estructura["hoja"]]
    assert hoja._images == []


# ---------------------------------------------------------------------------
# claves_obligatorias (backlog #12, PR 1 of a stacked-to-main chain; spec
# `listado-reportes` "Percent Avance Per Card"; design D1). Public so
# `reportes/listado.py` derives the same denominator `_validar_completitud`
# already uses internally, never a second obligatorio enumeration.
# ---------------------------------------------------------------------------


def test_claves_obligatorias_coincide_con_validar_completitud(definicion_valida):
    """Task 1.1 RED: on empty `valores`, `claves_obligatorias(estructura)`
    must return exactly the same ids `_validar_completitud` reports as
    missing (`ValoresIncompletos.faltantes`) — `definicion_valida`'s single
    obligatorio node is the "turno" campo; its "p-01" item is not
    obligatorio, so it must never appear."""
    from tipos_reporte.generador import (
        ValoresIncompletos,
        _validar_completitud,
        claves_obligatorias,
    )

    estructura = definicion_valida()

    with pytest.raises(ValoresIncompletos) as excinfo:
        _validar_completitud(estructura, {})

    assert set(claves_obligatorias(estructura)) == set(excinfo.value.faltantes)
